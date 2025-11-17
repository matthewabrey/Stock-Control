from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field as PydanticField, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import openpyxl
import io


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class Field(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    name: str
    area: str
    crop_type: str
    variety: str
    available_grades: List[str] = []
    harvest_year: str = "2025"  # Default harvest year
    type: Optional[str] = None  # Type column for classification (Red, Brown, Special, etc.)

class FieldCreate(BaseModel):
    name: str
    area: str
    crop_type: str
    variety: str
    available_grades: List[str] = []
    harvest_year: str = "2025"
    type: Optional[str] = None

class DoorPosition(BaseModel):
    side: str  # "top", "bottom", "left", "right"
    position: float  # Position along that side (in meters)

class Shed(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    name: str
    width: float
    height: float
    description: Optional[str] = None
    doors: List[DoorPosition] = []
    order: Optional[int] = 9999  # Order from Excel, default to end if not set

class ShedCreate(BaseModel):
    name: str
    width: float
    height: float
    description: Optional[str] = None
    doors: List[DoorPosition] = []

class Zone(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    shed_id: str
    name: str
    x: float
    y: float
    width: float
    height: float
    total_quantity: float = 0
    max_capacity: int = 6

class ZoneCreate(BaseModel):
    shed_id: str
    name: str
    x: float
    y: float
    width: float
    height: float
    max_capacity: int = 6

class Fridge(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    shed_id: str
    name: str  # Will be "Fridge"
    x: float
    y: float
    width: float
    height: float

class FridgeCreate(BaseModel):
    shed_id: str
    name: str
    x: float
    y: float
    width: float
    height: float

class Door(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    shed_id: str
    name: str  # Will be "Door"
    x: float
    y: float
    width: float
    height: float

class DoorCreate(BaseModel):
    shed_id: str
    name: str
    x: float
    y: float
    width: float
    height: float

class StockIntake(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    field_id: str
    field_name: str
    zone_id: str
    shed_id: str
    quantity: float
    date: str
    grade: Optional[str] = None
    created_at: str = PydanticField(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class StockIntakeCreate(BaseModel):
    field_id: str
    field_name: str
    zone_id: str
    shed_id: str
    quantity: float
    date: str
    grade: Optional[str] = None

class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    from_zone_id: str
    to_zone_id: str
    from_shed_id: str
    to_shed_id: str
    quantity: float
    date: str
    employee_number: Optional[str] = None  # Who moved the stock
    field_id: Optional[str] = None  # What was moved
    field_name: Optional[str] = None
    grade: Optional[str] = None
    created_at: str = PydanticField(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class StockMovementCreate(BaseModel):
    from_zone_id: str
    to_zone_id: str
    from_shed_id: str
    to_shed_id: str
    quantity: float
    date: str
    employee_number: Optional[str] = None
    field_id: Optional[str] = None
    field_name: Optional[str] = None
    grade: Optional[str] = None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = PydanticField(default_factory=lambda: str(uuid.uuid4()))
    employee_number: str
    name: str
    stock_movement: str = "No"  # Yes/No
    admin_control: str = "NO"  # YES/NO
    qc: Optional[str] = "No"
    daily_check: Optional[str] = "No"
    workshop_control: Optional[str] = "No"
    operations: Optional[str] = "No"

class UserCreate(BaseModel):
    employee_number: str
    name: str
    stock_movement: str = "No"
    admin_control: str = "NO"
    qc: Optional[str] = "No"
    daily_check: Optional[str] = "No"
    workshop_control: Optional[str] = "No"
    operations: Optional[str] = "No"

class LoginRequest(BaseModel):
    employee_number: str


# Field Routes
@api_router.post("/fields", response_model=Field)
async def create_field(input: FieldCreate):
    field_obj = Field(**input.model_dump())
    doc = field_obj.model_dump()
    await db.fields.insert_one(doc)
    return field_obj

@api_router.get("/fields", response_model=List[Field])
async def get_fields(harvest_year: Optional[str] = None):
    query = {}
    if harvest_year:
        query["harvest_year"] = harvest_year
    fields = await db.fields.find(query, {"_id": 0}).to_list(length=None)
    return fields

@api_router.get("/harvest-years")
async def get_harvest_years():
    """Get list of available harvest years from fields"""
    fields = await db.fields.find({}, {"_id": 0, "harvest_year": 1}).to_list(length=None)
    years = list(set([f.get("harvest_year", "2025") for f in fields]))
    years.sort()
    return {"harvest_years": years}

@api_router.delete("/fields/{field_id}")
async def delete_field(field_id: str):
    result = await db.fields.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Field not found")
    return {"message": "Field deleted"}


# Shed Routes
@api_router.post("/sheds", response_model=Shed)
async def create_shed(input: ShedCreate):
    shed_obj = Shed(**input.model_dump())
    doc = shed_obj.model_dump()
    await db.sheds.insert_one(doc)
    return shed_obj

@api_router.get("/sheds", response_model=List[Shed])
async def get_sheds():
    sheds = await db.sheds.find({}, {"_id": 0}).to_list(1000)
    # Sort by order field (Excel sheet order), fallback to name if order doesn't exist
    sheds.sort(key=lambda x: (x.get('order', 9999), x.get('name', '')))
    return sheds

@api_router.get("/sheds/{shed_id}", response_model=Shed)
async def get_shed(shed_id: str):
    shed = await db.sheds.find_one({"id": shed_id}, {"_id": 0})
    if not shed:
        raise HTTPException(status_code=404, detail="Shed not found")
    return shed

@api_router.delete("/sheds/{shed_id}")
async def delete_shed(shed_id: str):
    result = await db.sheds.delete_one({"id": shed_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shed not found")
    # Delete all zones and fridges in this shed
    await db.zones.delete_many({"shed_id": shed_id})
    await db.fridges.delete_many({"shed_id": shed_id})
    return {"message": "Shed deleted"}


# Zone Routes
@api_router.post("/zones", response_model=Zone)
async def create_zone(input: ZoneCreate):
    zone_obj = Zone(**input.model_dump())
    doc = zone_obj.model_dump()
    await db.zones.insert_one(doc)
    return zone_obj

@api_router.get("/zones", response_model=List[Zone])
async def get_zones(shed_id: Optional[str] = None):
    query = {"shed_id": shed_id} if shed_id else {}
    zones = await db.zones.find(query, {"_id": 0}).to_list(length=None)
    return zones

@api_router.put("/zones/{zone_id}", response_model=Zone)
async def update_zone(zone_id: str, quantity: float):
    zone = await db.zones.find_one({"id": zone_id}, {"_id": 0})
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    
    await db.zones.update_one(
        {"id": zone_id},
        {"$set": {"total_quantity": quantity}}
    )
    
    zone["total_quantity"] = quantity
    return zone

@api_router.delete("/zones/{zone_id}")
async def delete_zone(zone_id: str):
    result = await db.zones.delete_one({"id": zone_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Zone not found")
    return {"message": "Zone deleted"}


# Fridge Routes
@api_router.post("/fridges", response_model=Fridge)
async def create_fridge(input: FridgeCreate):
    fridge_obj = Fridge(**input.model_dump())
    doc = fridge_obj.model_dump()
    await db.fridges.insert_one(doc)
    return fridge_obj

@api_router.get("/fridges", response_model=List[Fridge])
async def get_fridges(shed_id: Optional[str] = None):
    query = {"shed_id": shed_id} if shed_id else {}
    fridges = await db.fridges.find(query, {"_id": 0}).to_list(length=None)
    return fridges

@api_router.delete("/fridges/{fridge_id}")
async def delete_fridge(fridge_id: str):
    result = await db.fridges.delete_one({"id": fridge_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fridge not found")
    return {"message": "Fridge deleted"}


# Door Routes
@api_router.post("/doors", response_model=Door)
async def create_door(input: DoorCreate):
    door_obj = Door(**input.model_dump())
    doc = door_obj.model_dump()
    await db.doors.insert_one(doc)
    return door_obj

@api_router.get("/doors", response_model=List[Door])
async def get_doors(shed_id: Optional[str] = None):
    query = {"shed_id": shed_id} if shed_id else {}
    doors = await db.doors.find(query, {"_id": 0}).to_list(length=None)
    return doors

@api_router.delete("/doors/{door_id}")
async def delete_door(door_id: str):
    result = await db.doors.delete_one({"id": door_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Door not found")
    return {"message": "Door deleted"}


# Stock Intake Routes
@api_router.post("/stock-intakes", response_model=StockIntake)
async def create_stock_intake(input: StockIntakeCreate):
    intake_obj = StockIntake(**input.model_dump())
    doc = intake_obj.model_dump()
    await db.stock_intakes.insert_one(doc)
    
    # Update zone quantity
    zone = await db.zones.find_one({"id": input.zone_id})
    if zone:
        new_quantity = zone.get("total_quantity", 0) + input.quantity
        await db.zones.update_one(
            {"id": input.zone_id},
            {"$set": {"total_quantity": new_quantity}}
        )
    
    return intake_obj

@api_router.get("/stock-intakes", response_model=List[StockIntake])
async def get_stock_intakes():
    intakes = await db.stock_intakes.find({}, {"_id": 0}).to_list(None)  # No limit
    return intakes

@api_router.get("/stock-intakes/zone/{zone_id}", response_model=List[StockIntake])
async def get_zone_stock_intakes(zone_id: str):
    intakes = await db.stock_intakes.find({"zone_id": zone_id}, {"_id": 0}).to_list(1000)
    return intakes

@api_router.put("/stock-intakes/{intake_id}", response_model=StockIntake)
async def update_stock_intake(intake_id: str, input: StockIntakeCreate):
    # Get existing intake
    existing = await db.stock_intakes.find_one({"id": intake_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Stock intake not found")
    
    # Update the intake
    intake_obj = StockIntake(**input.model_dump())
    intake_obj.id = intake_id  # Keep same ID
    doc = intake_obj.model_dump()
    await db.stock_intakes.update_one({"id": intake_id}, {"$set": doc})
    
    return intake_obj

@api_router.delete("/stock-intakes/{intake_id}")
async def delete_stock_intake(intake_id: str):
    result = await db.stock_intakes.delete_one({"id": intake_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock intake not found")
    return {"message": "Stock intake deleted"}


# Stock Movement Routes
@api_router.post("/stock-movements", response_model=StockMovement)
async def create_stock_movement(input: StockMovementCreate):
    # Check if source zone has enough quantity
    from_zone = await db.zones.find_one({"id": input.from_zone_id})
    if not from_zone:
        raise HTTPException(status_code=404, detail="Source zone not found")
    
    if from_zone.get("total_quantity", 0) < input.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock in source zone")
    
    # Create movement record
    movement_obj = StockMovement(**input.model_dump())
    doc = movement_obj.model_dump()
    await db.stock_movements.insert_one(doc)
    
    # Update zone quantities
    await db.zones.update_one(
        {"id": input.from_zone_id},
        {"$inc": {"total_quantity": -input.quantity}}
    )
    
    to_zone = await db.zones.find_one({"id": input.to_zone_id})
    if to_zone:
        await db.zones.update_one(
            {"id": input.to_zone_id},
            {"$inc": {"total_quantity": input.quantity}}
        )
    
    return movement_obj

@api_router.post("/log-movement")
async def log_movement(input: StockMovementCreate):
    """Log a stock movement without validation or quantity updates (for tracking only)"""
    movement_obj = StockMovement(**input.model_dump())
    doc = movement_obj.model_dump()
    await db.stock_movements.insert_one(doc)
    return {"message": "Movement logged successfully"}

@api_router.get("/stock-movements", response_model=List[StockMovement])
async def get_stock_movements():
    movements = await db.stock_movements.find({}, {"_id": 0}).to_list(length=None)
    return movements


# User Management and Authentication Routes
@api_router.post("/login")
async def login(input: LoginRequest):
    """Login with employee number"""
    # Hardcoded admin access for employee 1234
    if input.employee_number == "1234":
        return {
            "id": str(uuid.uuid4()),
            "employee_number": "1234",
            "name": "Admin User",
            "stock_movement": "Yes",
            "admin_control": "YES",
            "qc": "Yes",
            "daily_check": "Yes",
            "workshop_control": "Yes",
            "operations": "Yes"
        }
    
    user = await db.users.find_one({"employee_number": input.employee_number}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Employee number not found")
    return user

@api_router.get("/users", response_model=List[User])
async def get_users():
    users = await db.users.find({}, {"_id": 0}).to_list(length=None)
    return users

@api_router.get("/users/{employee_number}", response_model=User)
async def get_user(employee_number: str):
    user = await db.users.find_one({"employee_number": employee_number}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@api_router.post("/upload-name-list")
async def upload_name_list(file: UploadFile = File(...)):
    """Upload and parse name list Excel file"""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        # Get the first sheet (assuming name list is in first sheet)
        ws = wb.active
        
        # Clear existing users
        await db.users.delete_many({})
        
        users_created = 0
        
        # Find header row (usually row 1)
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value
            if header:
                headers.append(str(header).strip())
        
        print(f"Found headers: {headers}")
        
        # Parse user data starting from row 2
        for row_idx in range(2, ws.max_row + 1):
            employee_number = ws.cell(row_idx, 1).value
            name = ws.cell(row_idx, 2).value
            
            if not employee_number or not name:
                continue
            
            # Parse columns based on headers
            user_doc = {
                "id": str(uuid.uuid4()),
                "employee_number": str(employee_number).strip(),
                "name": str(name).strip(),
                "qc": str(ws.cell(row_idx, 3).value or "No").strip(),
                "daily_check": str(ws.cell(row_idx, 4).value or "No").strip(),
                "stock_movement": str(ws.cell(row_idx, 5).value or "No").strip(),
                "workshop_control": str(ws.cell(row_idx, 6).value or "No").strip(),
                "admin_control": str(ws.cell(row_idx, 7).value or "NO").strip().upper(),
                "operations": str(ws.cell(row_idx, 8).value or "No").strip()
            }
            
            await db.users.insert_one(user_doc)
            users_created += 1
            print(f"Created user: {user_doc['employee_number']} - {user_doc['name']} (Admin: {user_doc['admin_control']})")
        
        return {
            "message": f"Name list uploaded successfully. {users_created} users created.",
            "users_created": users_created
        }
    
    except Exception as e:
        print(f"Error uploading name list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload name list: {str(e)}")


# Excel Upload for Fields and Store Plans
@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        fields_created = 0
        stores_created = 0
        zones_created = 0
        
        # Parse grade tables from "Grade Options Page" sheet
        grade_tables = {
            'onion': [],
            'onion_special': [],
            'maincrop': [],
            'salad': [],
            'carrot': []
        }
        
        if "Grade Options Page" in wb.sheetnames:
            ws_grades = wb["Grade Options Page"]
            print("=== Parsing Grade Options Page ===")
            
            # Row 1 contains headers (crop types)
            # Find which column corresponds to which crop type
            crop_columns = {}
            for col_idx in range(1, ws_grades.max_column + 1):
                header = ws_grades.cell(1, col_idx).value
                if header:
                    header_str = str(header).strip().lower()
                    print(f"Column {col_idx}: '{header}'")
                    
                    if 'onion' in header_str and 'special' in header_str:
                        crop_columns['onion_special'] = col_idx
                    elif 'onion' in header_str:
                        crop_columns['onion'] = col_idx
                    elif 'maincrop' in header_str or 'main crop' in header_str:
                        crop_columns['maincrop'] = col_idx
                    elif 'salad' in header_str:
                        crop_columns['salad'] = col_idx
                    elif 'carrot' in header_str:
                        crop_columns['carrot'] = col_idx
            
            print(f"Found crop columns: {crop_columns}")
            
            # Read grades from row 2 onwards for each crop type
            for crop_type, col_idx in crop_columns.items():
                grades = []
                for row_idx in range(2, ws_grades.max_row + 1):
                    grade_val = ws_grades.cell(row_idx, col_idx).value
                    if grade_val:
                        grade_str = str(grade_val).strip()
                        if grade_str:
                            grades.append(grade_str)
                
                grade_tables[crop_type] = grades
                print(f"{crop_type}: {len(grades)} grades - {grades[:3]}...")
            
            print(f"Final grade tables: {grade_tables}")
        else:
            print("Warning: 'Grade Options Page' sheet not found")
        
        # Parse harvest year sheets for fields
        harvest_sheets = []
        
        # Check for field data sheets (various possible names)
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower()
            if any(keyword in sheet_lower for keyword in ["master harvest", "master harevst", "master cropping", "front page", "fields"]):
                harvest_sheets.append(sheet_name)
        
        # If no recognized sheets, skip field import
        if not harvest_sheets:
            print("Warning: No field sheets found (looking for 'Master Harvest', 'Master Cropping', 'FRONT PAGE', etc.)")
        
        # Clear existing fields
        await db.fields.delete_many({})
        
        for sheet_name in harvest_sheets:
            ws = wb[sheet_name]
            
            print(f"\n=== Processing {sheet_name} ===")
            
            # Detect column layout by checking row 3 or row 4 for headers
            # Master Harvest 25: Row 3 has headers, data starts row 4, columns C-G
            # Master Harvest 26: Row 4 has headers, data starts row 5, columns D-H (or more if Year column exists)
            
            farm_col = 3  # Default: Column C
            field_col = 4  # Default: Column D
            area_col = 5  # Default: Column E
            crop_col = 6  # Default: Column F
            variety_col = 7  # Default: Column G
            type_col = 8  # Default: Column H (Type column)
            year_col = None  # Will be detected if exists
            start_row = 4  # Default data start row
            
            # Check if row 4 has header values (indicates Master Harvest 26 format)
            row4_field = ws.cell(4, 5).value  # Check column E in row 4
            if row4_field and str(row4_field).lower() in ['field', 'farm']:
                # Master Harvest 26 format: columns shifted right, headers in row 4
                farm_col = 4  # Column D
                field_col = 5  # Column E
                area_col = 6  # Column F
                crop_col = 7  # Column G
                variety_col = 8  # Column H
                type_col = 9  # Column I (Type column)
                start_row = 5  # Data starts row 5
                
                # Check for Year column (could be column J or beyond)
                for col_idx in range(10, ws.max_column + 1):
                    header_val = ws.cell(4, col_idx).value
                    if header_val and 'year' in str(header_val).lower():
                        year_col = col_idx
                        break
                
                print(f"Detected Harvest 26 format: columns D-H, Type=I, starting row 5, year_col={year_col}")
            else:
                # Check for Year column in row 3 (Master Harvest 25 format)
                for col_idx in range(9, ws.max_column + 1):
                    header_val = ws.cell(3, col_idx).value
                    if header_val and 'year' in str(header_val).lower():
                        year_col = col_idx
                        break
                
                print(f"Detected Harvest 25 format: columns C-G, Type=H, starting row 4, year_col={year_col}")
            
            # Parse fields from data start row onwards
            for row_idx in range(start_row, ws.max_row + 1):
                farm = ws.cell(row_idx, farm_col).value
                field_name = ws.cell(row_idx, field_col).value
                area = ws.cell(row_idx, area_col).value
                crop = ws.cell(row_idx, crop_col).value
                variety_excel = ws.cell(row_idx, variety_col).value  # Column 7 = Classification (Red/Brown/Special)
                type_excel = ws.cell(row_idx, type_col).value  # Column 8 = Actual variety name
                
                # Read year from column if it exists, otherwise use sheet name
                if year_col:
                    year_value = ws.cell(row_idx, year_col).value
                    harvest_year = str(year_value) if year_value else "2025"
                else:
                    # Fallback: extract from sheet name
                    if "25" in sheet_name:
                        harvest_year = "2025"
                    elif "26" in sheet_name:
                        harvest_year = "2026"
                    else:
                        harvest_year = "2025"
                
                if not farm or not field_name:
                    continue
                
                # Assign grades based on crop type from parsed tables
                grades = []
                crop_str = str(crop).lower() if crop else ""
                
                # For grade matching, use the classification from variety_excel (Column 7)
                classification_str = str(variety_excel).lower() if variety_excel else ""
                
                # Match crop type to grade table
                if 'onion' in crop_str:
                    # Check if it's a special onion variety
                    if 'special' in classification_str or 'shallot' in classification_str or 'special' in crop_str.lower():
                        grades = grade_tables.get('onion_special', grade_tables.get('onion', []))
                    else:
                        grades = grade_tables.get('onion', [])
                elif 'maincrop' in crop_str or 'main crop' in crop_str or 'potato' in crop_str:
                    grades = grade_tables.get('maincrop', [])
                elif 'salad' in crop_str:
                    grades = grade_tables.get('salad', [])
                elif 'carrot' in crop_str:
                    grades = grade_tables.get('carrot', [])
                
                # If no grades found, add a default
                if not grades:
                    grades = ['Whole Crop']
                
                full_field_name = f"{farm} - {field_name}"
                area_str = f"{area} Acres" if area else "N/A"
                
                # IMPORTANT: Excel columns are:
                # - Column 7 (variety_excel) = Classification (Red, Brown, Special)
                # - Column 8 (type_excel) = Actual variety name (Figaro, Hybound, etc.)
                # So we store: variety = type_excel (actual variety name)
                #              type = variety_excel (classification)
                field_doc = {
                    "id": str(uuid.uuid4()),
                    "name": full_field_name,
                    "area": area_str,
                    "crop_type": str(crop) if crop else "Unknown",
                    "variety": str(type_excel) if type_excel else "Unknown",  # Column 8 = variety name
                    "available_grades": grades,
                    "harvest_year": harvest_year,
                    "type": str(variety_excel) if variety_excel else None  # Column 7 = classification (Red/Brown/Special)
                }
                await db.fields.insert_one(field_doc)
                fields_created += 1
                print(f"  Created field: {full_field_name} (Harvest {harvest_year})")
        
        # Parse Store Sheets (each sheet = one store)
        # Skip field sheets, Grade Options Page, and other non-store sheets
        skip_sheets = ["FRONT PAGE", "Master Harvest 25", "Master Harevst 26", "Master Harvest 26", "Master Cropping", "Grade Options Page", "Sheet1", "Sheet2", "Sheet3"]
        
        print("=== Processing Store Sheets ===")
        print(f"All sheets in workbook: {wb.sheetnames}")
        print(f"Skip list: {skip_sheets}")
        
        sheet_order = 0  # Track order of sheets
        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                print(f"Skipping sheet: '{sheet_name}' (in skip list)")
                continue
            
            store_name = sheet_name.strip()
            
            # Check if store already exists
            existing_shed = await db.sheds.find_one({"name": store_name})
            if existing_shed:
                print(f"Store '{store_name}' already exists in database, skipping...")
                continue
            
            print(f"Processing new store: '{store_name}'")
            sheet_order += 1  # Increment order for each processed sheet
            
            ws = wb[sheet_name]
            
            # First, scan for storage type indicator (Box or Bulk)
            # Look for cells containing "Box" or "Bulk" keywords
            storage_type = "box"  # Default to box storage
            for row_idx in range(1, min(20, ws.max_row + 1)):  # Check first 20 rows
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value:
                        cell_str = str(cell.value).lower()
                        if 'bulk' in cell_str:
                            storage_type = "bulk"
                            print("  Detected BULK storage type")
                            break
                        elif 'box' in cell_str:
                            storage_type = "box"
                            print("  Detected BOX storage type")
                            break
                if storage_type == "bulk":
                    break
            
            # Find all zones, fridges, and doors - scan the entire sheet
            zone_positions = []  # Will store (row, col, capacity, width, height)
            fridge_positions = []  # Will store (row, col, width, height) for fridges
            door_positions = []  # Will store (row, col, width, height) for doors (blue cells)
            max_col = 0
            max_row = 0
            min_col = float('inf')
            min_row = float('inf')
            processed_cells = set()  # Track cells we've already processed
            
            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    # Skip if we've already processed this cell as part of a merged range
                    if (row_idx, col_idx) in processed_cells:
                        continue
                    
                    cell = ws.cell(row_idx, col_idx)
                    
                    # Handle merged cells properly - get the actual value from top-left cell
                    cell_value = None
                    cell_width = 1
                    cell_height = 1
                    is_merged = False
                    
                    # Check if this cell is part of a merged range
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            # Only process if this is the top-left cell of the merge
                            if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                                cell_value = cell.value
                                cell_height = merged_range.max_row - merged_range.min_row + 1
                                cell_width = merged_range.max_col - merged_range.min_col + 1
                                
                                # Mark all cells in this merged range as processed
                                for r in range(merged_range.min_row, merged_range.max_row + 1):
                                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                                        processed_cells.add((r, c))
                            else:
                                # This is NOT the top-left, skip it
                                break
                            break
                    
                    # If not part of a merged range, it's a regular cell
                    if not is_merged:
                        cell_value = cell.value
                    
                    if cell_value is not None:
                        # Safely convert to string
                        try:
                            cell_str = str(cell_value).strip()
                        except Exception as e:
                            print(f"  Warning: Could not convert cell value to string: {e}")
                            continue
                        
                        # Check for DOOR markers (blue cells with "Door" text)
                        if 'door' in cell_str.lower():
                            # Check if cell has blue fill
                            cell_fill = cell.fill
                            is_blue = False
                            if cell_fill and cell_fill.start_color:
                                # Check for blue color (hex: 0000FF, 0070C0, etc.)
                                color_value = cell_fill.start_color.rgb if hasattr(cell_fill.start_color, 'rgb') else None
                                if color_value:
                                    # Blue variants: FF0000FF, 0000FF, 000070C0, FF0070C0, etc.
                                    color_str = str(color_value).upper()
                                    # Check for various blue shades
                                    if ('0000FF' in color_str or '0070C0' in color_str or 
                                        '4472C4' in color_str or '5B9BD5' in color_str):
                                        is_blue = True
                            
                            if is_blue:
                                print(f"  Found DOOR at row={row_idx}, col={col_idx}, size={cell_width}x{cell_height}")
                                door_positions.append((row_idx, col_idx, cell_width, cell_height))
                                max_col = max(max_col, col_idx + cell_width - 1)
                                max_row = max(max_row, row_idx + cell_height - 1)
                                min_col = min(min_col, col_idx)
                                min_row = min(min_row, row_idx)
                                continue  # Don't process as zone
                        
                        # Check for FRIDGE markers (yellow cells with "Fridge" text)
                        if 'fridge' in cell_str.lower():
                            # Check if cell has yellow fill
                            cell_fill = cell.fill
                            is_yellow = False
                            if cell_fill and cell_fill.start_color:
                                # Check for yellow color (hex: FFFF00 or similar)
                                color_value = cell_fill.start_color.rgb if hasattr(cell_fill.start_color, 'rgb') else None
                                if color_value:
                                    # Yellow variants: FFFFFF00, FFFF00, 00FFFF00, etc.
                                    color_str = str(color_value).upper()
                                    if 'FFFF00' in color_str or (color_str.endswith('FFFF00')):
                                        is_yellow = True
                            
                            if is_yellow:
                                print(f"  Found FRIDGE at row={row_idx}, col={col_idx}, size={cell_width}x{cell_height}")
                                fridge_positions.append((row_idx, col_idx, cell_width, cell_height))
                                max_col = max(max_col, col_idx + cell_width - 1)
                                max_row = max(max_row, row_idx + cell_height - 1)
                                min_col = min(min_col, col_idx)
                                min_row = min(min_row, row_idx)
                                continue  # Don't process as zone
                        
                        # Check for bulk storage (tonnage like "175t", "200t", etc.)
                        if cell_str.lower().endswith('t') and len(cell_str) > 1:
                            try:
                                # Extract the number before 't'
                                tonnage = int(cell_str[:-1])
                                # Create ONE zone for the entire merged cell
                                zone_positions.append((row_idx, col_idx, tonnage, cell_width, cell_height))
                                max_col = max(max_col, col_idx + cell_width - 1)
                                max_row = max(max_row, row_idx + cell_height - 1)
                                min_col = min(min_col, col_idx)
                                min_row = min(min_row, row_idx)
                            except (ValueError, TypeError) as e:
                                print(f"  Warning: Could not parse tonnage '{cell_str}': {e}")
                                pass
                        # Check for numeric capacity (box storage like "5", "6", "7", "8", etc.)
                        else:
                            try:
                                capacity = int(cell_str)
                                # Only accept reasonable capacity numbers (1-50)
                                if 1 <= capacity <= 50:
                                    # Create ONE zone for the entire merged cell
                                    zone_positions.append((row_idx, col_idx, capacity, cell_width, cell_height))
                                    max_col = max(max_col, col_idx + cell_width - 1)
                                    max_row = max(max_row, row_idx + cell_height - 1)
                                    min_col = min(min_col, col_idx)
                                    min_row = min(min_row, row_idx)
                            except (ValueError, TypeError) as e:
                                print(f"  Warning: Could not parse capacity '{cell_str}': {e}")
                                pass
            
            if not zone_positions:
                print(f"No zones found in {store_name}, skipping...")
                continue
            
            print(f"Store {store_name}: Found {len(zone_positions)} zones")
            print(f"  Bounds: rows {min_row}-{max_row}, cols {min_col}-{max_col}")
            
            # Debug: Show zone distribution by row
            from collections import defaultdict
            zones_by_row = defaultdict(list)
            for row_idx, col_idx, capacity, cell_width, cell_height in zone_positions:
                zones_by_row[row_idx].append((col_idx, cell_width, cell_height, capacity))
            
            print("  Zone distribution by row:")
            for row in sorted(zones_by_row.keys())[:5]:  # Show first 5 rows
                zones = zones_by_row[row]
                widths = [z[1] for z in zones]
                print(f"    Row {row}: {len(zones)} zones, widths: {widths[:10]}{'...' if len(widths) > 10 else ''}")
            
            # Calculate zone, fridge, and door positions
            # Group zones, fridges, and doors by row for proper x-position calculation
            zones_by_row = {}
            for row_idx, col_idx, capacity, cell_width, cell_height in zone_positions:
                if row_idx not in zones_by_row:
                    zones_by_row[row_idx] = []
                zones_by_row[row_idx].append((col_idx, capacity, cell_width, cell_height, 'zone'))
            
            # Add fridges to the same row structure for position calculation
            for row_idx, col_idx, cell_width, cell_height in fridge_positions:
                if row_idx not in zones_by_row:
                    zones_by_row[row_idx] = []
                zones_by_row[row_idx].append((col_idx, 0, cell_width, cell_height, 'fridge'))  # capacity=0 for fridges
            
            # Add doors to the same row structure for position calculation
            for row_idx, col_idx, cell_width, cell_height in door_positions:
                if row_idx not in zones_by_row:
                    zones_by_row[row_idx] = []
                zones_by_row[row_idx].append((col_idx, 0, cell_width, cell_height, 'door'))  # capacity=0 for doors
            
            # Sort items in each row by column
            for row_idx in zones_by_row:
                zones_by_row[row_idx].sort(key=lambda x: x[0])  # Sort by col_idx
            
            # Calculate x positions for each zone based on its row
            # This handles mixed merged/unmerged cells properly
            zone_x_positions = {}  # {(row_idx, col_idx): x_position}
            max_width_per_row = {}  # Track max width of each row
            
            for row_idx in sorted(zones_by_row.keys()):
                row_items = zones_by_row[row_idx]
                current_x = 0
                prev_col_idx = min_col - 1
                
                for col_idx, capacity, cell_width, cell_height, item_type in row_items:
                    # Check if there are empty columns between previous item and this one
                    if col_idx > prev_col_idx + 1:
                        # Add gaps for empty columns
                        gap_cols = col_idx - (prev_col_idx + 1)
                        current_x += gap_cols * 2  # 2m per empty column
                    
                    # Store position for this zone or fridge
                    zone_x_positions[(row_idx, col_idx)] = current_x
                    
                    # Calculate width and advance current_x
                    if storage_type == "bulk":
                        item_width = 8 * cell_width
                    else:
                        item_width = 2 * cell_width
                    
                    current_x += item_width
                    prev_col_idx = col_idx + cell_width - 1  # Last column occupied by this item
                
                max_width_per_row[row_idx] = current_x
            
            # Calculate total store dimensions
            # Use the widest row as the store width
            store_width = max(max_width_per_row.values()) + 2  # Add 2m buffer
            store_height = (max_row - min_row + 1) * 2
            
            print(f"  Store dimensions: width={store_width}m, height={store_height}m")
            print(f"  Row widths: {[(row, width) for row, width in max_width_per_row.items()][:5]}...")
            
            # Detect doors - look for cells containing "DOOR" text (both inside and outside grid)
            doors = []
            
            # Process doors that were found inside the grid
            for door_row, door_col in door_positions:
                door_side = None
                door_position = 0
                
                # Determine which edge this door is closest to
                # Check if it's on the boundary of the zone grid
                if door_col == min_col:
                    # Left edge
                    door_side = 'left'
                    door_position = (door_row - min_row) * 2
                elif door_col == max_col:
                    # Right edge
                    door_side = 'right'
                    door_position = (door_row - min_row) * 2
                elif door_row == min_row:
                    # Top edge
                    door_side = 'top'
                    # Use simple calculation for door position
                    door_position = (door_col - min_col) * 2
                elif door_row == max_row:
                    # Bottom edge
                    door_side = 'bottom'
                    # Use simple calculation for door position
                    door_position = (door_col - min_col) * 2
                else:
                    # Door is in the middle of grid, use closest edge
                    # For now, default to right side
                    door_side = 'right'
                    door_position = (door_row - min_row) * 2
                
                if door_side:
                    door_dict = {"side": door_side, "position": door_position}
                    if door_dict not in doors:
                        doors.append(door_dict)
                        print(f"  Added door: {door_side} at {door_position}m (from grid cell {openpyxl.utils.get_column_letter(door_col)}{door_row})")
            
            # Also check for doors OUTSIDE the grid (original logic)
            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value and 'door' in str(cell.value).lower():
                        # Skip if already processed (was in door_positions)
                        if (row_idx, col_idx) in door_positions:
                            continue
                        
                        # Determine which side this door is on relative to the zone area
                        door_side = None
                        door_position = 0
                        
                        # Top: row is above zone area
                        if row_idx < min_row and col_idx >= min_col and col_idx <= max_col:
                            door_side = 'top'
                            # Use simple calculation for door position
                            door_position = (col_idx - min_col) * 2
                        # Bottom: row is below zone area
                        elif row_idx > max_row and col_idx >= min_col and col_idx <= max_col:
                            door_side = 'bottom'
                            # Use simple calculation for door position
                            door_position = (col_idx - min_col) * 2
                        # Left: column is left of zone area
                        elif col_idx < min_col and row_idx >= min_row and row_idx <= max_row:
                            door_side = 'left'
                            door_position = (row_idx - min_row) * 2
                        # Right: column is right of zone area
                        elif col_idx > max_col and row_idx >= min_row and row_idx <= max_row:
                            door_side = 'right'
                            door_position = (row_idx - min_row) * 2
                        
                        if door_side:
                            door_dict = {"side": door_side, "position": door_position}
                            if door_dict not in doors:
                                doors.append(door_dict)
                                print(f"  Found door: {door_side} at {door_position}m (cell {openpyxl.utils.get_column_letter(col_idx)}{row_idx})")
            
            # Create shed
            shed_id = str(uuid.uuid4())
            shed_doc = {
                "id": shed_id,
                "name": store_name,
                "width": store_width,
                "height": store_height,
                "description": f"Imported from Excel - {len(zone_positions)} zones",
                "doors": doors,
                "order": sheet_order  # Preserve Excel sheet order
            }
            await db.sheds.insert_one(shed_doc)
            stores_created += 1
            
            # Create zones
            # Use the zone_x_positions we calculated above
            for row_idx, col_idx, capacity, cell_width, cell_height in zone_positions:
                # Get position from our calculated positions
                zone_x = zone_x_positions.get((row_idx, col_idx), 0)
                zone_y = (row_idx - min_row) * 2
                
                # Generate zone name (column letter + row number)
                col_letter = openpyxl.utils.get_column_letter(col_idx - min_col + 1)
                zone_name = f"{col_letter}{row_idx - min_row + 1}"
                
                # Use merged cell dimensions to determine zone size
                if storage_type == "bulk":
                    base_width = 8  # Bulk storage base width
                else:
                    base_width = 2  # Box storage base width
                
                # Scale zone dimensions based on merged cell size
                zone_width = base_width * cell_width
                zone_height = 2 * cell_height
                
                zone_doc = {
                    "id": str(uuid.uuid4()),
                    "shed_id": shed_id,
                    "name": zone_name,
                    "x": zone_x,
                    "y": zone_y,
                    "width": zone_width,
                    "height": zone_height,
                    "total_quantity": 0,
                    "max_capacity": capacity  # Use the capacity from the cell (6 for boxes, tonnage for bulk)
                }
                await db.zones.insert_one(zone_doc)
                zones_created += 1
            
            # Create fridges
            fridges_created = 0
            for fridge_row, fridge_col, fridge_cell_width, fridge_cell_height in fridge_positions:
                # Calculate fridge position using same logic as zones
                # For fridges, we need to calculate x position based on its row
                fridge_x = zone_x_positions.get((fridge_row, fridge_col), 0)
                fridge_y = (fridge_row - min_row) * 2
                
                # Fridge dimensions (same as zones)
                if storage_type == "bulk":
                    base_width = 8
                else:
                    base_width = 2
                
                fridge_width = base_width * fridge_cell_width
                fridge_height = 2 * fridge_cell_height
                
                fridge_doc = {
                    "id": str(uuid.uuid4()),
                    "shed_id": shed_id,
                    "name": "Fridge",
                    "x": fridge_x,
                    "y": fridge_y,
                    "width": fridge_width,
                    "height": fridge_height
                }
                await db.fridges.insert_one(fridge_doc)
                fridges_created += 1
                print(f"  Created fridge at ({fridge_x}, {fridge_y}) with size {fridge_width}x{fridge_height}")
            
            if fridges_created > 0:
                print(f"Store {store_name}: Created {fridges_created} fridges")
        
        return {
            "message": "Excel uploaded successfully",
            "fields_created": fields_created,
            "stores_created": stores_created,
            "zones_created": zones_created
        }
    
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# Database integrity check endpoint
@api_router.get("/database-integrity")
async def check_database_integrity():
    """Check database integrity and report any inconsistencies"""
    try:
        issues = []
        stats = {}
        
        # Count documents
        sheds_count = await db.sheds.count_documents({})
        zones_count = await db.zones.count_documents({})
        intakes_count = await db.stock_intakes.count_documents({})
        fields_count = await db.fields.count_documents({})
        
        stats["sheds"] = sheds_count
        stats["zones"] = zones_count
        stats["stock_intakes"] = intakes_count
        stats["fields"] = fields_count
        
        # Check for orphaned zones (zones whose shed_id doesn't exist)
        zones = await db.zones.find({}).to_list(length=None)
        shed_ids = set([s["id"] for s in await db.sheds.find({}).to_list(length=None)])
        
        orphaned_zones = []
        for zone in zones:
            if zone["shed_id"] not in shed_ids:
                orphaned_zones.append({"zone_id": zone["id"], "zone_name": zone["name"], "invalid_shed_id": zone["shed_id"]})
        
        if orphaned_zones:
            issues.append({
                "type": "orphaned_zones",
                "count": len(orphaned_zones),
                "message": f"Found {len(orphaned_zones)} zones with invalid shed_id",
                "examples": orphaned_zones[:5]
            })
        
        # Check for stock intakes with invalid zone_id or shed_id
        intakes = await db.stock_intakes.find({}).to_list(length=None)
        zone_ids = set([z["id"] for z in zones])
        
        invalid_intakes = []
        for intake in intakes:
            if intake["zone_id"] not in zone_ids:
                invalid_intakes.append({
                    "intake_id": intake["id"],
                    "invalid_zone_id": intake["zone_id"],
                    "shed_id": intake["shed_id"]
                })
            elif intake["shed_id"] not in shed_ids:
                invalid_intakes.append({
                    "intake_id": intake["id"],
                    "zone_id": intake["zone_id"],
                    "invalid_shed_id": intake["shed_id"]
                })
        
        if invalid_intakes:
            issues.append({
                "type": "invalid_stock_intakes",
                "count": len(invalid_intakes),
                "message": f"Found {len(invalid_intakes)} stock intakes with invalid zone_id or shed_id",
                "examples": invalid_intakes[:5]
            })
        
        # Check for zone quantity mismatches
        quantity_mismatches = []
        for zone in zones:
            zone_intakes = [i for i in intakes if i["zone_id"] == zone["id"]]
            expected_qty = sum(i["quantity"] for i in zone_intakes)
            actual_qty = zone.get("total_quantity", 0)
            
            if abs(expected_qty - actual_qty) > 0.01:  # Allow small floating point differences
                quantity_mismatches.append({
                    "zone_id": zone["id"],
                    "zone_name": zone["name"],
                    "shed_id": zone["shed_id"],
                    "expected_quantity": expected_qty,
                    "actual_quantity": actual_qty,
                    "difference": expected_qty - actual_qty
                })
        
        if quantity_mismatches:
            issues.append({
                "type": "quantity_mismatches",
                "count": len(quantity_mismatches),
                "message": f"Found {len(quantity_mismatches)} zones where total_quantity doesn't match sum of intakes",
                "examples": quantity_mismatches[:10]
            })
        
        return {
            "status": "healthy" if len(issues) == 0 else "issues_found",
            "stats": stats,
            "issues": issues,
            "message": "No issues found" if len(issues) == 0 else f"Found {len(issues)} types of issues"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking integrity: {str(e)}")


# Clear all data endpoint
@api_router.delete("/clear-all-data")
async def clear_all_data():
    """Clear all data from the database"""
    try:
        # Delete all documents from each collection
        await db.fields.delete_many({})
        await db.sheds.delete_many({})
        await db.zones.delete_many({})
        await db.fridges.delete_many({})
        await db.stock_intakes.delete_many({})
        await db.stock_movements.delete_many({})
        
        return {
            "message": "All data cleared successfully",
            "collections_cleared": ["fields", "sheds", "zones", "fridges", "stock_intakes", "stock_movements"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error clearing data: {str(e)}")


@api_router.delete("/clear-stores")
async def clear_stores():
    """Clear all stock from sheds/zones, but keep the shed and zone structure intact"""
    try:
        # Delete all stock intakes and movements
        await db.stock_intakes.delete_many({})
        await db.stock_movements.delete_many({})
        
        # Reset all zone quantities to 0
        await db.zones.update_many({}, {"$set": {"total_quantity": 0}})
        
        return {
            "message": "All stock cleared successfully. Sheds and zones preserved.",
            "actions": ["stock_intakes cleared", "stock_movements cleared", "zone quantities reset to 0"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error clearing stock: {str(e)}")


@api_router.get("/export-excel")
async def export_excel():
    """Export all data to Excel file"""
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Export Fields
        ws_fields = wb.create_sheet("Fields")
        ws_fields.append(["ID", "Name", "Area", "Crop Type", "Variety", "Harvest Year", "Available Grades"])
        
        fields = await db.fields.find({}).to_list(length=None)
        for field in fields:
            ws_fields.append([
                field.get('id', ''),
                field.get('name', ''),
                field.get('area', ''),
                field.get('crop_type', ''),
                field.get('variety', ''),
                field.get('harvest_year', ''),
                ', '.join(field.get('available_grades', []))
            ])
        
        # Export Sheds
        ws_sheds = wb.create_sheet("Sheds")
        ws_sheds.append(["ID", "Name", "Width", "Height", "Description"])
        
        sheds = await db.sheds.find({}).to_list(length=None)
        for shed in sheds:
            ws_sheds.append([
                shed.get('id', ''),
                shed.get('name', ''),
                shed.get('width', 0),
                shed.get('height', 0),
                shed.get('description', '')
            ])
        
        # Export Zones
        ws_zones = wb.create_sheet("Zones")
        ws_zones.append(["ID", "Shed ID", "Name", "X", "Y", "Width", "Height", "Total Quantity", "Max Capacity"])
        
        zones = await db.zones.find({}).to_list(length=None)
        for zone in zones:
            ws_zones.append([
                zone.get('id', ''),
                zone.get('shed_id', ''),
                zone.get('name', ''),
                zone.get('x', 0),
                zone.get('y', 0),
                zone.get('width', 0),
                zone.get('height', 0),
                zone.get('total_quantity', 0),
                zone.get('max_capacity', 6)
            ])
        
        # Export Stock Intakes
        ws_intakes = wb.create_sheet("Stock Intakes")
        ws_intakes.append(["ID", "Field ID", "Field Name", "Zone ID", "Shed ID", "Quantity", "Grade", "Date"])
        
        intakes = await db.stock_intakes.find({}).to_list(length=None)
        for intake in intakes:
            ws_intakes.append([
                intake.get('id', ''),
                intake.get('field_id', ''),
                intake.get('field_name', ''),
                intake.get('zone_id', ''),
                intake.get('shed_id', ''),
                intake.get('quantity', 0),
                intake.get('grade', ''),
                intake.get('date', '')
            ])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stock-control-export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to Excel: {str(e)}")


# Root route
@api_router.get("/")
async def root():
    return {"message": "Stock Control API"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()