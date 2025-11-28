#!/usr/bin/env python3
"""
Backend Integrity Testing for Stock Control System After Data Reset
Tests database integrity, Excel upload flow, and data consistency
"""

import requests
import json
import io
import openpyxl
from datetime import datetime

# Backend URL from environment
BACKEND_URL = "https://shed-mapper.preview.emergentagent.com/api"

class StockControlIntegrityTester:
    def __init__(self):
        self.base_url = BACKEND_URL
        self.session = requests.Session()
        self.test_results = []
        
    def log_test(self, test_name, success, message, details=None):
        """Log test results"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details:
            print(f"   Details: {details}")
    
    def create_test_excel_with_grade_options(self):
        """Create a comprehensive test Excel file with Grade Options Page"""
        wb = openpyxl.Workbook()
        
        # Create Grade Options Page sheet first
        grade_ws = wb.create_sheet("Grade Options Page")
        wb.remove(wb.active)  # Remove default sheet
        
        # Set up grade table headers in row 1
        grade_ws['A1'] = "Onion"
        grade_ws['B1'] = "Onion Special"
        grade_ws['C1'] = "Maincrop Potato"
        grade_ws['D1'] = "Salad Potato"
        grade_ws['E1'] = "Carrot"
        
        # Add grade values starting from row 2
        onion_grades = ["O1", "O2", "O3", "O4"]
        onion_special_grades = ["OS1", "OS2", "OS3"]
        maincrop_grades = ["MC1", "MC2", "MC3", "MC4", "MC5", "MC6", "MC7"]
        salad_grades = ["SP1", "SP2", "SP3"]
        carrot_grades = ["C1", "C2", "C3"]
        
        for i, grade in enumerate(onion_grades, start=2):
            grade_ws[f'A{i}'] = grade
        for i, grade in enumerate(onion_special_grades, start=2):
            grade_ws[f'B{i}'] = grade
        for i, grade in enumerate(maincrop_grades, start=2):
            grade_ws[f'C{i}'] = grade
        for i, grade in enumerate(salad_grades, start=2):
            grade_ws[f'D{i}'] = grade
        for i, grade in enumerate(carrot_grades, start=2):
            grade_ws[f'E{i}'] = grade
        
        # Create FRONT PAGE sheet
        front_ws = wb.create_sheet("FRONT PAGE")
        
        # Field data headers (row 3)
        front_ws['C3'] = "Farm"
        front_ws['D3'] = "Field"
        front_ws['E3'] = "Area"
        front_ws['F3'] = "Crop"
        front_ws['G3'] = "Variety"
        
        # Test field data (starting from row 4)
        test_fields = [
            ("Greenfield Farm", "Field A", "25", "Onion", "Red Baron"),
            ("Hillside Farm", "Field B", "30", "Maincrop Potato", "Maris Piper"),
            ("Valley Farm", "Field C", "15", "Salad Potato", "Charlotte"),
            ("Riverside Farm", "Field D", "20", "Onion", "Shallot Special"),
            ("Mountain Farm", "Field E", "18", "Carrot", "Nantes")
        ]
        
        for i, (farm, field, area, crop, variety) in enumerate(test_fields, start=4):
            front_ws[f'C{i}'] = farm
            front_ws[f'D{i}'] = field
            front_ws[f'E{i}'] = area
            front_ws[f'F{i}'] = crop
            front_ws[f'G{i}'] = variety
        
        # Create store sheets
        # Store 1: Mixed storage
        store1_ws = wb.create_sheet("Main Storage")
        store1_ws['B2'] = "6"  # Box storage
        store1_ws['C2'] = "6"
        store1_ws['D2'] = "6"
        store1_ws['B3'] = "175t"  # Bulk storage
        store1_ws['C3'] = "200t"
        store1_ws['A1'] = "DOOR"  # Door marker
        
        # Store 2: Box storage only
        store2_ws = wb.create_sheet("Grader Shed")
        store2_ws['A1'] = "6"
        store2_ws['B1'] = "6"
        store2_ws['C1'] = "6"
        store2_ws['A2'] = "6"
        store2_ws['B2'] = "6"
        store2_ws['C2'] = "6"
        store2_ws['D1'] = "DOOR"
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def test_database_integrity_clean_state(self):
        """Test database integrity check on clean state"""
        try:
            response = self.session.get(f"{self.base_url}/database-integrity")
            
            if response.status_code != 200:
                self.log_test("Database Integrity Check", False, f"API returned status {response.status_code}")
                return False
            
            integrity_data = response.json()
            
            # Check status
            if integrity_data.get('status') != 'healthy':
                self.log_test("Database Integrity Check", False, f"Expected 'healthy' status, got '{integrity_data.get('status')}'")
                return False
            
            # Check stats - all should be 0
            stats = integrity_data.get('stats', {})
            expected_counts = {'sheds': 0, 'zones': 0, 'stock_intakes': 0, 'fields': 0}
            
            for key, expected in expected_counts.items():
                actual = stats.get(key, -1)
                if actual != expected:
                    self.log_test("Database Integrity Check", False, f"Expected {key}={expected}, got {actual}")
                    return False
            
            # Check no issues reported
            issues = integrity_data.get('issues', [])
            if issues:
                self.log_test("Database Integrity Check", False, f"Found {len(issues)} issues in clean database", str(issues))
                return False
            
            self.log_test("Database Integrity Check", True, "Clean database state verified - all counts at 0, no issues")
            return True
            
        except Exception as e:
            self.log_test("Database Integrity Check", False, f"Exception: {str(e)}")
            return False
    
    def test_basic_endpoints_empty(self):
        """Test that basic endpoints return empty arrays"""
        endpoints = [
            ("fields", "/fields"),
            ("sheds", "/sheds"),
            ("zones", "/zones"),
            ("stock-intakes", "/stock-intakes")
        ]
        
        all_passed = True
        
        for name, endpoint in endpoints:
            try:
                response = self.session.get(f"{self.base_url}{endpoint}")
                
                if response.status_code != 200:
                    self.log_test(f"Empty {name} endpoint", False, f"Status {response.status_code}")
                    all_passed = False
                    continue
                
                data = response.json()
                if not isinstance(data, list) or len(data) != 0:
                    self.log_test(f"Empty {name} endpoint", False, f"Expected empty array, got {type(data)} with {len(data) if isinstance(data, list) else 'N/A'} items")
                    all_passed = False
                    continue
                
                self.log_test(f"Empty {name} endpoint", True, "Returns empty array as expected")
                
            except Exception as e:
                self.log_test(f"Empty {name} endpoint", False, f"Exception: {str(e)}")
                all_passed = False
        
        return all_passed
    
    def test_excel_upload_flow(self):
        """Test complete Excel upload flow"""
        try:
            # Create and upload Excel file
            excel_data = self.create_test_excel_with_grade_options()
            
            files = {'file': ('test_complete.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{self.base_url}/upload-excel", files=files)
            
            if response.status_code != 200:
                self.log_test("Excel Upload Flow", False, f"Upload failed with status {response.status_code}", response.text)
                return False
            
            result = response.json()
            fields_created = result.get('fields_created', 0)
            stores_created = result.get('stores_created', 0)
            zones_created = result.get('zones_created', 0)
            
            # Verify reasonable numbers were created
            if fields_created < 3:
                self.log_test("Excel Upload Flow", False, f"Expected at least 3 fields, got {fields_created}")
                return False
            
            if stores_created < 2:
                self.log_test("Excel Upload Flow", False, f"Expected at least 2 stores, got {stores_created}")
                return False
            
            if zones_created < 5:
                self.log_test("Excel Upload Flow", False, f"Expected at least 5 zones, got {zones_created}")
                return False
            
            self.log_test("Excel Upload Flow", True, f"Successfully created {fields_created} fields, {stores_created} stores, {zones_created} zones")
            return True
            
        except Exception as e:
            self.log_test("Excel Upload Flow", False, f"Exception: {str(e)}")
            return False
    
    def test_post_upload_integrity(self):
        """Test database integrity after Excel upload"""
        try:
            response = self.session.get(f"{self.base_url}/database-integrity")
            
            if response.status_code != 200:
                self.log_test("Post-Upload Integrity", False, f"API returned status {response.status_code}")
                return False
            
            integrity_data = response.json()
            
            # Should still be healthy with no data integrity issues
            if integrity_data.get('status') != 'healthy':
                issues = integrity_data.get('issues', [])
                self.log_test("Post-Upload Integrity", False, f"Status '{integrity_data.get('status')}' with issues", str(issues))
                return False
            
            # Check that we have data now
            stats = integrity_data.get('stats', {})
            if stats.get('fields', 0) == 0 or stats.get('sheds', 0) == 0 or stats.get('zones', 0) == 0:
                self.log_test("Post-Upload Integrity", False, f"Expected non-zero counts after upload: {stats}")
                return False
            
            # Verify no orphaned records or integrity issues
            issues = integrity_data.get('issues', [])
            if issues:
                self.log_test("Post-Upload Integrity", False, f"Found {len(issues)} integrity issues after upload", str(issues))
                return False
            
            self.log_test("Post-Upload Integrity", True, f"Database healthy after upload with stats: {stats}")
            return True
            
        except Exception as e:
            self.log_test("Post-Upload Integrity", False, f"Exception: {str(e)}")
            return False
    
    def test_stock_intake_creation_and_integrity(self):
        """Test stock intake creation and verify zone quantity updates"""
        try:
            # Get available fields, sheds, and zones
            fields_response = self.session.get(f"{self.base_url}/fields")
            sheds_response = self.session.get(f"{self.base_url}/sheds")
            
            if fields_response.status_code != 200 or sheds_response.status_code != 200:
                self.log_test("Stock Intake Setup", False, "Failed to get fields or sheds")
                return False
            
            fields = fields_response.json()
            sheds = sheds_response.json()
            
            if not fields or not sheds:
                self.log_test("Stock Intake Setup", False, "No fields or sheds available")
                return False
            
            # Get zones for first shed
            shed = sheds[0]
            zones_response = self.session.get(f"{self.base_url}/zones?shed_id={shed['id']}")
            
            if zones_response.status_code != 200:
                self.log_test("Stock Intake Setup", False, "Failed to get zones")
                return False
            
            zones = zones_response.json()
            if not zones:
                self.log_test("Stock Intake Setup", False, "No zones available")
                return False
            
            field = fields[0]
            zone = zones[0]
            
            # Record initial zone quantity
            initial_quantity = zone.get('total_quantity', 0)
            
            # Create stock intake
            intake_quantity = 75.5
            intake_data = {
                "field_id": field['id'],
                "field_name": field['name'],
                "zone_id": zone['id'],
                "shed_id": shed['id'],
                "quantity": intake_quantity,
                "date": "2024-01-20",
                "grade": field.get('available_grades', ['Whole Crop'])[0]
            }
            
            response = self.session.post(f"{self.base_url}/stock-intakes", json=intake_data)
            
            if response.status_code != 200:
                self.log_test("Stock Intake Creation", False, f"Failed to create stock intake, status: {response.status_code}")
                return False
            
            # Verify zone quantity was updated
            zones_response = self.session.get(f"{self.base_url}/zones?shed_id={shed['id']}")
            updated_zones = zones_response.json()
            updated_zone = next((z for z in updated_zones if z['id'] == zone['id']), None)
            
            if not updated_zone:
                self.log_test("Stock Intake Creation", False, "Zone not found after stock intake")
                return False
            
            expected_quantity = initial_quantity + intake_quantity
            actual_quantity = updated_zone.get('total_quantity', 0)
            
            if abs(actual_quantity - expected_quantity) > 0.01:
                self.log_test("Stock Intake Creation", False, f"Zone quantity mismatch. Expected: {expected_quantity}, Got: {actual_quantity}")
                return False
            
            self.log_test("Stock Intake Creation", True, f"Stock intake created and zone quantity updated correctly ({initial_quantity} + {intake_quantity} = {actual_quantity})")
            return True
            
        except Exception as e:
            self.log_test("Stock Intake Creation", False, f"Exception: {str(e)}")
            return False
    
    def test_final_integrity_check(self):
        """Final integrity check after all operations"""
        try:
            response = self.session.get(f"{self.base_url}/database-integrity")
            
            if response.status_code != 200:
                self.log_test("Final Integrity Check", False, f"API returned status {response.status_code}")
                return False
            
            integrity_data = response.json()
            
            # Should still be healthy
            if integrity_data.get('status') != 'healthy':
                issues = integrity_data.get('issues', [])
                self.log_test("Final Integrity Check", False, f"Status '{integrity_data.get('status')}' with issues", str(issues))
                return False
            
            # Check for any quantity mismatches
            issues = integrity_data.get('issues', [])
            quantity_issues = [issue for issue in issues if issue.get('type') == 'quantity_mismatches']
            
            if quantity_issues:
                self.log_test("Final Integrity Check", False, f"Found quantity mismatches", str(quantity_issues))
                return False
            
            stats = integrity_data.get('stats', {})
            self.log_test("Final Integrity Check", True, f"Database remains healthy after all operations. Stats: {stats}")
            return True
            
        except Exception as e:
            self.log_test("Final Integrity Check", False, f"Exception: {str(e)}")
            return False
    
    def test_clear_all_data(self):
        """Test clearing all data and verify clean state"""
        try:
            # Clear all data
            response = self.session.delete(f"{self.base_url}/clear-all-data")
            
            if response.status_code != 200:
                self.log_test("Clear All Data", False, f"Clear operation failed with status {response.status_code}")
                return False
            
            result = response.json()
            collections_cleared = result.get('collections_cleared', [])
            expected_collections = ['fields', 'sheds', 'zones', 'stock_intakes', 'stock_movements']
            
            for collection in expected_collections:
                if collection not in collections_cleared:
                    self.log_test("Clear All Data", False, f"Collection '{collection}' not in cleared list")
                    return False
            
            # Verify integrity check shows clean state
            integrity_response = self.session.get(f"{self.base_url}/database-integrity")
            
            if integrity_response.status_code != 200:
                self.log_test("Clear All Data", False, "Failed to check integrity after clear")
                return False
            
            integrity_data = integrity_response.json()
            
            if integrity_data.get('status') != 'healthy':
                self.log_test("Clear All Data", False, f"Database not healthy after clear: {integrity_data.get('status')}")
                return False
            
            stats = integrity_data.get('stats', {})
            for key in ['sheds', 'zones', 'stock_intakes', 'fields']:
                if stats.get(key, -1) != 0:
                    self.log_test("Clear All Data", False, f"Collection '{key}' not empty after clear: {stats.get(key)}")
                    return False
            
            self.log_test("Clear All Data", True, "All data cleared successfully, database returned to clean state")
            return True
            
        except Exception as e:
            self.log_test("Clear All Data", False, f"Exception: {str(e)}")
            return False
    
    def run_integrity_tests(self):
        """Run all integrity tests in sequence"""
        print(f"🧪 Starting Stock Control Backend Integrity Tests After Data Reset")
        print(f"🔗 Backend URL: {self.base_url}")
        print("=" * 70)
        
        # Test sequence as specified in the review request
        tests = [
            ("1. Database Integrity Check (Clean State)", self.test_database_integrity_clean_state),
            ("2. Basic Endpoint Health (Empty Arrays)", self.test_basic_endpoints_empty),
            ("3. Excel Upload Flow", self.test_excel_upload_flow),
            ("4. Post-Upload Integrity Check", self.test_post_upload_integrity),
            ("5. Stock Intake Creation & Zone Updates", self.test_stock_intake_creation_and_integrity),
            ("6. Final Integrity Check", self.test_final_integrity_check),
            ("7. Clear All Data", self.test_clear_all_data)
        ]
        
        passed = 0
        total = len(tests)
        
        for test_name, test_func in tests:
            print(f"\n🔍 Running: {test_name}")
            if test_func():
                passed += 1
            else:
                # If a critical test fails, we might want to continue but note it
                print(f"   ⚠️  Test failed but continuing with remaining tests...")
        
        print("\n" + "=" * 70)
        print(f"📊 Integrity Test Results: {passed}/{total} tests passed")
        
        # Show failed tests
        failed_tests = [result for result in self.test_results if not result['success']]
        if failed_tests:
            print("\n❌ Failed Tests:")
            for test in failed_tests:
                print(f"   • {test['test']}: {test['message']}")
                if test['details']:
                    print(f"     Details: {test['details']}")
        else:
            print("\n✅ All integrity tests passed! Database is functioning correctly after reset.")
        
        return passed == total

if __name__ == "__main__":
    tester = StockControlIntegrityTester()
    success = tester.run_integrity_tests()
    exit(0 if success else 1)