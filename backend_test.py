#!/usr/bin/env python3
"""
Backend API Testing for Stock Control System
Tests Excel upload with grade parsing, Field API, and basic CRUD operations
"""

import requests
import json
import io
import openpyxl
from datetime import datetime

# Backend URL from environment
BACKEND_URL = "https://cropflow-7.preview.emergentagent.com/api"

class StockControlTester:
    def __init__(self):
        self.base_url = BACKEND_URL
        self.session = requests.Session()
        self.test_results = []
        
    def log_test(self, test_name, success, message, details=None):
        """Log test results"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details:
            print(f"   Details: {details}")
    
    def create_test_excel_with_type(self):
        """Create a test Excel file with Type column for testing Type column integration"""
        wb = openpyxl.Workbook()
        
        # Create Grade Options Page
        ws_grades = wb.create_sheet("Grade Options Page")
        
        # Add grade tables
        ws_grades['A1'] = "Onion"
        ws_grades['A2'] = "50/60"
        ws_grades['A3'] = "70/80"
        ws_grades['A4'] = "O Whole Crop"
        
        ws_grades['B1'] = "Onion Special"
        ws_grades['B2'] = "Special Grade 1"
        ws_grades['B3'] = "Special Grade 2"
        
        ws_grades['C1'] = "Maincrop Potato"
        ws_grades['C2'] = "40/50"
        ws_grades['C3'] = "50/60"
        ws_grades['C4'] = "70/80"
        ws_grades['C5'] = "80+"
        
        # Create Master Harvest 25 sheet with Type column
        ws = wb.create_sheet("Master Harvest 25")
        
        # Headers in row 3 (Master Harvest 25 format)
        ws['C3'] = "Farm"
        ws['D3'] = "Field"
        ws['E3'] = "Area"
        ws['F3'] = "Crop"
        ws['G3'] = "Variety"
        ws['H3'] = "Type"  # Type column (Column H for Harvest 25)
        
        # Test field data with Type values (starting from row 4)
        test_fields = [
            ("Greenfield Farm", "Field A", "25", "Onion", "Red Baron", "Red"),
            ("Hillside Farm", "Field B", "30", "Onion", "Brown Variety", "Brown"),
            ("Valley Farm", "Field C", "15", "Onion", "Special Shallot", "Special"),
            ("Riverside Farm", "Field D", "20", "Maincrop Potato", "King Edward", "Brown"),
            ("Westside Farm", "Field E", "18", "Onion", "White Variety", None)  # No Type value
        ]
        
        for i, (farm, field, area, crop, variety, type_val) in enumerate(test_fields, start=4):
            ws[f'C{i}'] = farm
            ws[f'D{i}'] = field
            ws[f'E{i}'] = area
            ws[f'F{i}'] = crop
            ws[f'G{i}'] = variety
            if type_val:
                ws[f'H{i}'] = type_val
        
        # Create a simple store sheet with fridges
        store_ws = wb.create_sheet("Test Store")
        # Add some zones
        store_ws['B2'] = "6"  # Box storage
        store_ws['C2'] = "6"
        store_ws['B3'] = "175t"  # Bulk storage
        store_ws['A1'] = "DOOR"  # Door marker
        
        # Add a fridge with yellow fill and "Fridge" text
        fridge_cell = store_ws['D2']
        fridge_cell.value = "Fridge"
        # Set yellow fill color
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fridge_cell.fill = yellow_fill
        
        # Remove default sheet
        wb.remove(wb['Sheet'])
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def create_test_excel(self):
        """Create a test Excel file with grade tables and field data (legacy method)"""
        return self.create_test_excel_with_type()
    
    def test_api_health(self):
        """Test if API is accessible"""
        try:
            response = self.session.get(f"{self.base_url}/")
            if response.status_code == 200:
                self.log_test("API Health Check", True, "API is accessible")
                return True
            else:
                self.log_test("API Health Check", False, f"API returned status {response.status_code}")
                return False
        except Exception as e:
            self.log_test("API Health Check", False, f"Failed to connect: {str(e)}")
            return False
    
    def test_excel_upload_grade_parsing(self):
        """Test Excel upload with grade table parsing"""
        try:
            # Create test Excel file
            excel_data = self.create_test_excel()
            
            # Upload Excel file
            files = {'file': ('test_fields.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{self.base_url}/upload-excel", files=files)
            
            if response.status_code == 200:
                result = response.json()
                fields_created = result.get('fields_created', 0)
                stores_created = result.get('stores_created', 0)
                zones_created = result.get('zones_created', 0)
                
                self.log_test(
                    "Excel Upload", 
                    True, 
                    f"Successfully uploaded Excel file",
                    f"Created {fields_created} fields, {stores_created} stores, {zones_created} zones"
                )
                return True
            else:
                self.log_test("Excel Upload", False, f"Upload failed with status {response.status_code}", response.text)
                return False
                
        except Exception as e:
            self.log_test("Excel Upload", False, f"Exception during upload: {str(e)}")
            return False
    
    def test_fields_api_grades(self):
        """Test Fields API and verify grade assignment"""
        try:
            response = self.session.get(f"{self.base_url}/fields")
            
            if response.status_code != 200:
                self.log_test("Fields API", False, f"Failed to get fields, status: {response.status_code}")
                return False
            
            fields = response.json()
            
            if not fields:
                self.log_test("Fields API", False, "No fields found after Excel upload")
                return False
            
            # Test grade assignment logic
            grade_tests = []
            for field in fields:
                field_name = field.get('name', 'Unknown')
                crop_type = field.get('crop_type', 'Unknown')
                available_grades = field.get('available_grades', [])
                
                # Check if at least one grade is available (Whole Crop or crop-specific grades)
                if not available_grades:
                    grade_tests.append(f"❌ {field_name}: No grades available")
                    continue
                
                # Check crop-specific grades
                if 'Onion' in crop_type:
                    expected_grades = ['O1', 'O2', 'O3', 'O4']
                    missing_grades = [g for g in expected_grades if g not in available_grades]
                    if missing_grades:
                        grade_tests.append(f"❌ {field_name} (Onion): Missing grades {missing_grades}")
                    else:
                        grade_tests.append(f"✅ {field_name} (Onion): Correct grades {[g for g in available_grades if g.startswith('O')]}")
                
                elif 'Maincrop' in crop_type:
                    expected_grades = ['MC1', 'MC2', 'MC3', 'MC4', 'MC5', 'MC6', 'MC7']
                    missing_grades = [g for g in expected_grades if g not in available_grades]
                    if missing_grades:
                        grade_tests.append(f"❌ {field_name} (Maincrop): Missing grades {missing_grades}")
                    else:
                        grade_tests.append(f"✅ {field_name} (Maincrop): Correct grades {[g for g in available_grades if g.startswith('MC')]}")
                
                elif 'Salad' in crop_type:
                    expected_grades = ['SP1', 'SP2', 'SP3']
                    missing_grades = [g for g in expected_grades if g not in available_grades]
                    if missing_grades:
                        grade_tests.append(f"❌ {field_name} (Salad): Missing grades {missing_grades}")
                    else:
                        grade_tests.append(f"✅ {field_name} (Salad): Correct grades {[g for g in available_grades if g.startswith('SP')]}")
            
            # Check if any grade assignment failed
            failed_tests = [test for test in grade_tests if test.startswith('❌')]
            
            if failed_tests:
                self.log_test(
                    "Grade Assignment", 
                    False, 
                    f"Grade assignment issues found",
                    "\n".join(grade_tests)
                )
                return False
            else:
                self.log_test(
                    "Grade Assignment", 
                    True, 
                    f"All {len(fields)} fields have correct grades",
                    "\n".join(grade_tests)
                )
                return True
                
        except Exception as e:
            self.log_test("Fields API", False, f"Exception: {str(e)}")
            return False
    
    def test_shed_crud(self):
        """Test Shed CRUD operations"""
        try:
            # Create shed
            shed_data = {
                "name": "Test Shed API",
                "width": 50.0,
                "height": 30.0,
                "description": "Test shed for API testing",
                "doors": [{"side": "top", "position": 10.0}]
            }
            
            response = self.session.post(f"{self.base_url}/sheds", json=shed_data)
            if response.status_code != 200:
                self.log_test("Shed Creation", False, f"Failed to create shed, status: {response.status_code}")
                return False
            
            shed = response.json()
            shed_id = shed.get('id')
            
            # Get sheds
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Shed Retrieval", False, f"Failed to get sheds, status: {response.status_code}")
                return False
            
            sheds = response.json()
            test_shed = next((s for s in sheds if s.get('id') == shed_id), None)
            
            if not test_shed:
                self.log_test("Shed Retrieval", False, "Created shed not found in list")
                return False
            
            # Delete shed
            response = self.session.delete(f"{self.base_url}/sheds/{shed_id}")
            if response.status_code != 200:
                self.log_test("Shed Deletion", False, f"Failed to delete shed, status: {response.status_code}")
                return False
            
            self.log_test("Shed CRUD", True, "All shed operations successful")
            return True
            
        except Exception as e:
            self.log_test("Shed CRUD", False, f"Exception: {str(e)}")
            return False
    
    def test_zone_crud(self):
        """Test Zone CRUD operations"""
        try:
            # First create a shed for the zone
            shed_data = {
                "name": "Zone Test Shed",
                "width": 20.0,
                "height": 15.0
            }
            
            response = self.session.post(f"{self.base_url}/sheds", json=shed_data)
            if response.status_code != 200:
                self.log_test("Zone Test Setup", False, "Failed to create test shed for zones")
                return False
            
            shed = response.json()
            shed_id = shed.get('id')
            
            # Create zone
            zone_data = {
                "shed_id": shed_id,
                "name": "A1",
                "x": 5.0,
                "y": 5.0,
                "width": 2.0,
                "height": 2.0,
                "max_capacity": 6
            }
            
            response = self.session.post(f"{self.base_url}/zones", json=zone_data)
            if response.status_code != 200:
                self.log_test("Zone Creation", False, f"Failed to create zone, status: {response.status_code}")
                return False
            
            zone = response.json()
            zone_id = zone.get('id')
            
            # Get zones
            response = self.session.get(f"{self.base_url}/zones?shed_id={shed_id}")
            if response.status_code != 200:
                self.log_test("Zone Retrieval", False, f"Failed to get zones, status: {response.status_code}")
                return False
            
            zones = response.json()
            if not zones or zones[0].get('id') != zone_id:
                self.log_test("Zone Retrieval", False, "Created zone not found")
                return False
            
            # Update zone quantity
            response = self.session.put(f"{self.base_url}/zones/{zone_id}?quantity=100")
            if response.status_code != 200:
                self.log_test("Zone Update", False, f"Failed to update zone, status: {response.status_code}")
                return False
            
            # Clean up
            self.session.delete(f"{self.base_url}/zones/{zone_id}")
            self.session.delete(f"{self.base_url}/sheds/{shed_id}")
            
            self.log_test("Zone CRUD", True, "All zone operations successful")
            return True
            
        except Exception as e:
            self.log_test("Zone CRUD", False, f"Exception: {str(e)}")
            return False
    
    def test_type_column_integration(self):
        """Test Type column parsing and storage from Excel"""
        try:
            # First clear all data
            response = self.session.delete(f"{self.base_url}/clear-all-data")
            if response.status_code != 200:
                self.log_test("Clear Data", False, f"Failed to clear data, status: {response.status_code}")
                return False
            
            # Create and upload Excel with Type column
            excel_data = self.create_test_excel_with_type()
            files = {'file': ('test_type_fields.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{self.base_url}/upload-excel", files=files)
            
            if response.status_code != 200:
                self.log_test("Type Column Upload", False, f"Upload failed with status {response.status_code}", response.text)
                return False
            
            # Get fields and verify Type column values
            response = self.session.get(f"{self.base_url}/fields")
            if response.status_code != 200:
                self.log_test("Type Column Fields API", False, f"Failed to get fields, status: {response.status_code}")
                return False
            
            fields = response.json()
            if not fields:
                self.log_test("Type Column Fields API", False, "No fields found after upload")
                return False
            
            # Expected Type values from our test data
            expected_types = {
                "Greenfield Farm - Field A": "Red",
                "Hillside Farm - Field B": "Brown", 
                "Valley Farm - Field C": "Special",
                "Riverside Farm - Field D": "Brown",
                "Westside Farm - Field E": None  # No Type value in Excel
            }
            
            type_tests = []
            for field in fields:
                field_name = field.get('name', 'Unknown')
                field_type = field.get('type')
                
                if field_name in expected_types:
                    expected_type = expected_types[field_name]
                    if field_type == expected_type:
                        type_tests.append(f"✅ {field_name}: Type='{field_type}' (correct)")
                    else:
                        type_tests.append(f"❌ {field_name}: Expected Type='{expected_type}', Got='{field_type}'")
                else:
                    type_tests.append(f"⚠️ {field_name}: Unexpected field (Type='{field_type}')")
            
            # Check if any Type tests failed
            failed_type_tests = [test for test in type_tests if test.startswith('❌')]
            
            if failed_type_tests:
                self.log_test(
                    "Type Column Integration", 
                    False, 
                    f"Type column parsing issues found",
                    "\n".join(type_tests)
                )
                return False
            else:
                self.log_test(
                    "Type Column Integration", 
                    True, 
                    f"All {len(fields)} fields have correct Type values",
                    "\n".join(type_tests)
                )
                return True
                
        except Exception as e:
            self.log_test("Type Column Integration", False, f"Exception: {str(e)}")
            return False
    
    def test_stock_intake_with_grade(self):
        """Test Stock Intake with grade field"""
        try:
            # Get existing fields to use for stock intake
            response = self.session.get(f"{self.base_url}/fields")
            if response.status_code != 200:
                self.log_test("Stock Intake Setup", False, "Failed to get fields for stock intake test")
                return False
            
            fields = response.json()
            if not fields:
                self.log_test("Stock Intake Setup", False, "No fields available for stock intake test")
                return False
            
            # Get existing sheds and zones
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Stock Intake Setup", False, "Failed to get sheds")
                return False
            
            sheds = response.json()
            if not sheds:
                self.log_test("Stock Intake Setup", False, "No sheds available for stock intake test")
                return False
            
            shed = sheds[0]
            shed_id = shed.get('id')
            
            # Get zones for this shed
            response = self.session.get(f"{self.base_url}/zones?shed_id={shed_id}")
            if response.status_code != 200:
                self.log_test("Stock Intake Setup", False, "Failed to get zones")
                return False
            
            zones = response.json()
            if not zones:
                self.log_test("Stock Intake Setup", False, "No zones available for stock intake test")
                return False
            
            field = fields[0]
            zone = zones[0]
            available_grades = field.get('available_grades', [])
            
            if not available_grades:
                self.log_test("Stock Intake Setup", False, "Field has no available grades")
                return False
            
            # Create stock intake with grade - use a specific grade if available
            selected_grade = available_grades[0]  # Default to first grade
            # Try to use a crop-specific grade if available
            for grade in available_grades:
                if grade.startswith(('O', 'MC', 'SP')) and grade != 'Whole Crop':
                    selected_grade = grade
                    break
            
            intake_data = {
                "field_id": field.get('id'),
                "field_name": field.get('name'),
                "zone_id": zone.get('id'),
                "shed_id": shed_id,
                "quantity": 50.0,
                "date": "2024-01-15",
                "grade": selected_grade
            }
            
            response = self.session.post(f"{self.base_url}/stock-intakes", json=intake_data)
            if response.status_code != 200:
                self.log_test("Stock Intake Creation", False, f"Failed to create stock intake, status: {response.status_code}")
                return False
            
            intake = response.json()
            
            # Verify grade was saved
            if intake.get('grade') != selected_grade:
                self.log_test("Stock Intake Grade", False, f"Grade not saved correctly. Expected: {selected_grade}, Got: {intake.get('grade')}")
                return False
            
            # Get stock intakes
            response = self.session.get(f"{self.base_url}/stock-intakes")
            if response.status_code != 200:
                self.log_test("Stock Intake Retrieval", False, f"Failed to get stock intakes, status: {response.status_code}")
                return False
            
            self.log_test("Stock Intake with Grade", True, f"Successfully created stock intake with grade '{selected_grade}'")
            return True
            
        except Exception as e:
            self.log_test("Stock Intake with Grade", False, f"Exception: {str(e)}")
            return False
    
    def test_fridge_api_endpoints(self):
        """Test Fridge API endpoints (POST, GET, DELETE)"""
        try:
            # First, ensure we have a shed to create fridges in
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Fridge API Setup", False, "Failed to get sheds for fridge test")
                return False
            
            sheds = response.json()
            if not sheds:
                # Create a test shed if none exist
                shed_data = {
                    "name": "Fridge Test Shed",
                    "width": 30.0,
                    "height": 20.0,
                    "description": "Test shed for fridge API testing"
                }
                response = self.session.post(f"{self.base_url}/sheds", json=shed_data)
                if response.status_code != 200:
                    self.log_test("Fridge API Setup", False, "Failed to create test shed for fridges")
                    return False
                shed = response.json()
            else:
                shed = sheds[0]
            
            shed_id = shed.get('id')
            
            # Test POST /api/fridges (create a fridge)
            fridge_data = {
                "shed_id": shed_id,
                "name": "Fridge",
                "x": 10.0,
                "y": 5.0,
                "width": 4.0,
                "height": 2.0
            }
            
            response = self.session.post(f"{self.base_url}/fridges", json=fridge_data)
            if response.status_code != 200:
                self.log_test("Fridge Creation", False, f"Failed to create fridge, status: {response.status_code}", response.text)
                return False
            
            fridge = response.json()
            fridge_id = fridge.get('id')
            
            # Verify fridge model has required fields
            required_fields = ['id', 'shed_id', 'name', 'x', 'y', 'width', 'height']
            missing_fields = [field for field in required_fields if field not in fridge]
            if missing_fields:
                self.log_test("Fridge Model Validation", False, f"Fridge missing required fields: {missing_fields}")
                return False
            
            # Verify field values
            if fridge.get('shed_id') != shed_id:
                self.log_test("Fridge Model Validation", False, f"Fridge shed_id mismatch. Expected: {shed_id}, Got: {fridge.get('shed_id')}")
                return False
            
            if fridge.get('name') != "Fridge":
                self.log_test("Fridge Model Validation", False, f"Fridge name mismatch. Expected: 'Fridge', Got: {fridge.get('name')}")
                return False
            
            # Test GET /api/fridges?shed_id=xxx (fetch fridges for a shed)
            response = self.session.get(f"{self.base_url}/fridges?shed_id={shed_id}")
            if response.status_code != 200:
                self.log_test("Fridge Retrieval by Shed", False, f"Failed to get fridges for shed, status: {response.status_code}")
                return False
            
            fridges = response.json()
            if not fridges:
                self.log_test("Fridge Retrieval by Shed", False, "No fridges found for shed after creation")
                return False
            
            # Verify our created fridge is in the list
            created_fridge = next((f for f in fridges if f.get('id') == fridge_id), None)
            if not created_fridge:
                self.log_test("Fridge Retrieval by Shed", False, "Created fridge not found in shed's fridge list")
                return False
            
            # Test GET /api/fridges (fetch all fridges)
            response = self.session.get(f"{self.base_url}/fridges")
            if response.status_code != 200:
                self.log_test("Fridge Retrieval All", False, f"Failed to get all fridges, status: {response.status_code}")
                return False
            
            all_fridges = response.json()
            created_fridge_all = next((f for f in all_fridges if f.get('id') == fridge_id), None)
            if not created_fridge_all:
                self.log_test("Fridge Retrieval All", False, "Created fridge not found in all fridges list")
                return False
            
            # Test DELETE /api/fridges/{fridge_id}
            response = self.session.delete(f"{self.base_url}/fridges/{fridge_id}")
            if response.status_code != 200:
                self.log_test("Fridge Deletion", False, f"Failed to delete fridge, status: {response.status_code}")
                return False
            
            # Verify fridge is deleted
            response = self.session.get(f"{self.base_url}/fridges?shed_id={shed_id}")
            if response.status_code == 200:
                remaining_fridges = response.json()
                deleted_fridge = next((f for f in remaining_fridges if f.get('id') == fridge_id), None)
                if deleted_fridge:
                    self.log_test("Fridge Deletion Verification", False, "Fridge still exists after deletion")
                    return False
            
            self.log_test("Fridge API Endpoints", True, "All fridge API operations successful (POST, GET, DELETE)")
            return True
            
        except Exception as e:
            self.log_test("Fridge API Endpoints", False, f"Exception: {str(e)}")
            return False
    
    def test_hardcoded_admin_login(self):
        """Test hardcoded admin login for employee 1234"""
        try:
            # Test login with hardcoded admin employee number "1234"
            login_data = {"employee_number": "1234"}
            
            response = self.session.post(f"{self.base_url}/login", json=login_data)
            if response.status_code != 200:
                self.log_test("Hardcoded Admin Login", False, f"Failed to login with employee 1234, status: {response.status_code}", response.text)
                return False
            
            admin_user = response.json()
            
            # Verify admin user has all required permissions
            expected_permissions = {
                "employee_number": "1234",
                "name": "Admin User",
                "stock_movement": "Yes",
                "admin_control": "YES",
                "qc": "Yes",
                "daily_check": "Yes",
                "workshop_control": "Yes",
                "operations": "Yes"
            }
            
            permission_tests = []
            for field, expected_value in expected_permissions.items():
                actual_value = admin_user.get(field)
                if actual_value == expected_value:
                    permission_tests.append(f"✅ {field}: '{actual_value}' (correct)")
                else:
                    permission_tests.append(f"❌ {field}: Expected '{expected_value}', Got '{actual_value}'")
            
            # Check if admin user has an ID
            if not admin_user.get('id'):
                permission_tests.append("❌ id: Missing ID field")
            else:
                permission_tests.append(f"✅ id: '{admin_user.get('id')}' (present)")
            
            # Check for any failed permission tests
            failed_permissions = [test for test in permission_tests if test.startswith('❌')]
            
            if failed_permissions:
                self.log_test(
                    "Hardcoded Admin Permissions", 
                    False, 
                    f"Admin permission issues found",
                    "\n".join(permission_tests)
                )
                return False
            
            # Test that normal employee numbers still work (if any users exist in DB)
            # First check if there are any users in the database
            response = self.session.get(f"{self.base_url}/users")
            if response.status_code == 200:
                users = response.json()
                if users:
                    # Try to login with a normal user
                    normal_user = users[0]
                    normal_employee_number = normal_user.get('employee_number')
                    
                    if normal_employee_number and normal_employee_number != "1234":
                        login_data = {"employee_number": normal_employee_number}
                        response = self.session.post(f"{self.base_url}/login", json=login_data)
                        
                        if response.status_code == 200:
                            normal_login_user = response.json()
                            self.log_test(
                                "Normal User Login", 
                                True, 
                                f"Normal user login still works for employee {normal_employee_number}"
                            )
                        else:
                            self.log_test(
                                "Normal User Login", 
                                False, 
                                f"Normal user login failed for employee {normal_employee_number}, status: {response.status_code}"
                            )
                            return False
                    else:
                        self.log_test(
                            "Normal User Login", 
                            True, 
                            "No normal users found to test (only admin 1234 exists)"
                        )
                else:
                    self.log_test(
                        "Normal User Login", 
                        True, 
                        "No users in database to test normal login"
                    )
            
            # Test invalid employee number
            login_data = {"employee_number": "9999"}
            response = self.session.post(f"{self.base_url}/login", json=login_data)
            if response.status_code == 404:
                self.log_test(
                    "Invalid Employee Login", 
                    True, 
                    "Invalid employee number correctly returns 404"
                )
            else:
                self.log_test(
                    "Invalid Employee Login", 
                    False, 
                    f"Invalid employee number should return 404, got {response.status_code}"
                )
                return False
            
            self.log_test(
                "Hardcoded Admin Login", 
                True, 
                "Admin login successful with all permissions",
                "\n".join(permission_tests)
            )
            return True
            
        except Exception as e:
            self.log_test("Hardcoded Admin Login", False, f"Exception: {str(e)}")
            return False
    
    def test_clear_data_with_fridges(self):
        """Test clear all data endpoint includes fridges collection"""
        try:
            # First create some test data including fridges
            # Create a shed
            shed_data = {
                "name": "Clear Test Shed",
                "width": 20.0,
                "height": 15.0
            }
            response = self.session.post(f"{self.base_url}/sheds", json=shed_data)
            if response.status_code != 200:
                self.log_test("Clear Data Setup", False, "Failed to create test shed")
                return False
            
            shed = response.json()
            shed_id = shed.get('id')
            
            # Create a fridge
            fridge_data = {
                "shed_id": shed_id,
                "name": "Fridge",
                "x": 5.0,
                "y": 5.0,
                "width": 2.0,
                "height": 2.0
            }
            response = self.session.post(f"{self.base_url}/fridges", json=fridge_data)
            if response.status_code != 200:
                self.log_test("Clear Data Setup", False, "Failed to create test fridge")
                return False
            
            # Verify fridge exists
            response = self.session.get(f"{self.base_url}/fridges")
            if response.status_code != 200:
                self.log_test("Clear Data Setup", False, "Failed to get fridges before clear")
                return False
            
            fridges_before = response.json()
            if not fridges_before:
                self.log_test("Clear Data Setup", False, "No fridges found before clear operation")
                return False
            
            # Test DELETE /api/clear-all-data
            response = self.session.delete(f"{self.base_url}/clear-all-data")
            if response.status_code != 200:
                self.log_test("Clear All Data", False, f"Failed to clear all data, status: {response.status_code}")
                return False
            
            clear_result = response.json()
            collections_cleared = clear_result.get('collections_cleared', [])
            
            # Verify fridges collection is in the cleared list
            if 'fridges' not in collections_cleared:
                self.log_test("Clear Data Fridges", False, f"Fridges collection not in cleared list: {collections_cleared}")
                return False
            
            # Verify fridges are actually deleted
            response = self.session.get(f"{self.base_url}/fridges")
            if response.status_code != 200:
                self.log_test("Clear Data Verification", False, "Failed to get fridges after clear")
                return False
            
            fridges_after = response.json()
            if fridges_after:
                self.log_test("Clear Data Verification", False, f"Fridges still exist after clear: {len(fridges_after)} fridges found")
                return False
            
            # Verify other collections are also cleared
            expected_collections = ["fields", "sheds", "zones", "fridges", "stock_intakes", "stock_movements"]
            missing_collections = [col for col in expected_collections if col not in collections_cleared]
            if missing_collections:
                self.log_test("Clear Data Collections", False, f"Missing collections from clear: {missing_collections}")
                return False
            
            self.log_test(
                "Clear All Data with Fridges", 
                True, 
                f"Successfully cleared all data including fridges collection",
                f"Cleared collections: {collections_cleared}"
            )
            return True
            
        except Exception as e:
            self.log_test("Clear All Data with Fridges", False, f"Exception: {str(e)}")
            return False
    
    def test_excel_fridge_parsing(self):
        """Test Excel parsing to detect yellow Fridge cells"""
        try:
            # Clear all data first
            response = self.session.delete(f"{self.base_url}/clear-all-data")
            if response.status_code != 200:
                self.log_test("Excel Fridge Setup", False, "Failed to clear data before fridge test")
                return False
            
            # Create and upload Excel with yellow fridge cells
            excel_data = self.create_test_excel_with_type()
            files = {'file': ('test_fridge_excel.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{self.base_url}/upload-excel", files=files)
            
            if response.status_code != 200:
                self.log_test("Excel Fridge Upload", False, f"Upload failed with status {response.status_code}", response.text)
                return False
            
            upload_result = response.json()
            
            # Get sheds to find the test store
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Excel Fridge Sheds", False, "Failed to get sheds after upload")
                return False
            
            sheds = response.json()
            test_store = next((s for s in sheds if s.get('name') == 'Test Store'), None)
            
            if not test_store:
                self.log_test("Excel Fridge Sheds", False, "Test Store not found after Excel upload")
                return False
            
            shed_id = test_store.get('id')
            
            # Get fridges for the test store
            response = self.session.get(f"{self.base_url}/fridges?shed_id={shed_id}")
            if response.status_code != 200:
                self.log_test("Excel Fridge Retrieval", False, f"Failed to get fridges for Test Store, status: {response.status_code}")
                return False
            
            fridges = response.json()
            
            if not fridges:
                self.log_test("Excel Fridge Parsing", False, "No fridges found after Excel upload with yellow Fridge cells")
                return False
            
            # Verify fridge properties
            fridge = fridges[0]
            fridge_tests = []
            
            if fridge.get('name') == 'Fridge':
                fridge_tests.append("✅ name: 'Fridge' (correct)")
            else:
                fridge_tests.append(f"❌ name: Expected 'Fridge', Got '{fridge.get('name')}'")
            
            if fridge.get('shed_id') == shed_id:
                fridge_tests.append(f"✅ shed_id: '{shed_id}' (correct)")
            else:
                fridge_tests.append(f"❌ shed_id: Expected '{shed_id}', Got '{fridge.get('shed_id')}'")
            
            # Check position and dimensions are present
            required_fields = ['x', 'y', 'width', 'height']
            for field in required_fields:
                value = fridge.get(field)
                if value is not None and isinstance(value, (int, float)):
                    fridge_tests.append(f"✅ {field}: {value} (present)")
                else:
                    fridge_tests.append(f"❌ {field}: Missing or invalid value '{value}'")
            
            # Check for any failed fridge tests
            failed_fridge_tests = [test for test in fridge_tests if test.startswith('❌')]
            
            if failed_fridge_tests:
                self.log_test(
                    "Excel Fridge Parsing", 
                    False, 
                    f"Fridge parsing issues found",
                    "\n".join(fridge_tests)
                )
                return False
            else:
                self.log_test(
                    "Excel Fridge Parsing", 
                    True, 
                    f"Successfully parsed {len(fridges)} fridge(s) from Excel yellow cells",
                    "\n".join(fridge_tests)
                )
                return True
                
        except Exception as e:
            self.log_test("Excel Fridge Parsing", False, f"Exception: {str(e)}")
            return False
    
    def test_review_request_workflow(self):
        """Test the specific review request workflow: clear data, upload new Excel with blue doors, verify doors and fridges"""
        try:
            print("\n🔍 EXECUTING REVIEW REQUEST WORKFLOW")
            print("=" * 50)
            
            # STEP 1: Clear all old data
            print("STEP 1: Clearing all old data...")
            response = self.session.delete(f"{self.base_url}/clear-all-data")
            if response.status_code != 200:
                self.log_test("Review Request - Clear Data", False, f"Failed to clear data, status: {response.status_code}")
                return False
            
            clear_result = response.json()
            collections_cleared = clear_result.get('collections_cleared', [])
            print(f"✅ Cleared collections: {collections_cleared}")
            
            # Verify fridges and doors are included in cleared collections
            if 'fridges' not in collections_cleared or 'doors' not in collections_cleared:
                self.log_test("Review Request - Clear Data", False, f"Fridges or doors not in cleared collections: {collections_cleared}")
                return False
            
            # STEP 2: Download and upload new Excel file
            print("\nSTEP 2: Downloading and uploading new Excel file...")
            
            # Download the new Excel file from the provided URL
            excel_url = "https://customer-assets.emergentagent.com/job_cropflow-7/artifacts/ydy00gnu_Stock%20Sheet%20Tables.xlsx"
            try:
                excel_response = requests.get(excel_url, timeout=30)
                if excel_response.status_code != 200:
                    self.log_test("Review Request - Download Excel", False, f"Failed to download Excel file, status: {excel_response.status_code}")
                    return False
                
                excel_data = excel_response.content
                print(f"✅ Downloaded Excel file ({len(excel_data)} bytes)")
                
            except Exception as e:
                self.log_test("Review Request - Download Excel", False, f"Exception downloading Excel: {str(e)}")
                return False
            
            # Upload the Excel file
            files = {'file': ('Stock_Sheet_Tables.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = self.session.post(f"{self.base_url}/upload-excel", files=files)
            
            if response.status_code != 200:
                self.log_test("Review Request - Upload Excel", False, f"Upload failed with status {response.status_code}", response.text)
                return False
            
            upload_result = response.json()
            fields_created = upload_result.get('fields_created', 0)
            stores_created = upload_result.get('stores_created', 0) 
            zones_created = upload_result.get('zones_created', 0)
            
            print(f"✅ Upload successful: {fields_created} fields, {stores_created} stores, {zones_created} zones created")
            
            # STEP 3: Verify doors and fridges in database
            print("\nSTEP 3: Verifying doors and fridges in database...")
            
            # Get all sheds
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Review Request - Get Sheds", False, f"Failed to get sheds, status: {response.status_code}")
                return False
            
            sheds = response.json()
            print(f"✅ Found {len(sheds)} sheds")
            
            # Get all doors
            response = self.session.get(f"{self.base_url}/doors")
            if response.status_code != 200:
                self.log_test("Review Request - Get Doors", False, f"Failed to get doors, status: {response.status_code}")
                return False
            
            all_doors = response.json()
            print(f"✅ Found {len(all_doors)} total doors")
            
            # Get all fridges
            response = self.session.get(f"{self.base_url}/fridges")
            if response.status_code != 200:
                self.log_test("Review Request - Get Fridges", False, f"Failed to get fridges, status: {response.status_code}")
                return False
            
            all_fridges = response.json()
            print(f"✅ Found {len(all_fridges)} total fridges")
            
            # STEP 4: Check doors and fridges for each shed
            print("\nSTEP 4: Checking doors and fridges for each shed...")
            
            shed_details = []
            for shed in sheds:
                shed_id = shed.get('id')
                shed_name = shed.get('name')
                
                # Get doors for this shed
                response = self.session.get(f"{self.base_url}/doors?shed_id={shed_id}")
                if response.status_code != 200:
                    print(f"❌ Failed to get doors for {shed_name}")
                    continue
                
                shed_doors = response.json()
                
                # Get fridges for this shed
                response = self.session.get(f"{self.base_url}/fridges?shed_id={shed_id}")
                if response.status_code != 200:
                    print(f"❌ Failed to get fridges for {shed_name}")
                    continue
                
                shed_fridges = response.json()
                
                shed_info = {
                    'name': shed_name,
                    'doors': len(shed_doors),
                    'fridges': len(shed_fridges),
                    'door_objects': shed_doors,
                    'fridge_objects': shed_fridges
                }
                shed_details.append(shed_info)
                
                print(f"🏢 {shed_name}: {len(shed_doors)} doors, {len(shed_fridges)} fridges")
            
            # STEP 5: Check backend logs (simulated by checking data integrity)
            print("\nSTEP 5: Checking data integrity...")
            
            # Verify door objects have correct structure
            door_validation = []
            for door in all_doors:
                required_fields = ['id', 'shed_id', 'name', 'x', 'y', 'width', 'height']
                missing_fields = [field for field in required_fields if field not in door]
                if missing_fields:
                    door_validation.append(f"❌ Door {door.get('id', 'unknown')} missing fields: {missing_fields}")
                else:
                    door_validation.append(f"✅ Door {door.get('id')[:8]}... has all required fields")
            
            # Verify fridge objects have correct structure
            fridge_validation = []
            for fridge in all_fridges:
                required_fields = ['id', 'shed_id', 'name', 'x', 'y', 'width', 'height']
                missing_fields = [field for field in required_fields if field not in fridge]
                if missing_fields:
                    fridge_validation.append(f"❌ Fridge {fridge.get('id', 'unknown')} missing fields: {missing_fields}")
                else:
                    fridge_validation.append(f"✅ Fridge {fridge.get('id')[:8]}... has all required fields")
            
            # Compile results
            results_summary = {
                'collections_cleared': collections_cleared,
                'upload_stats': {
                    'fields_created': fields_created,
                    'stores_created': stores_created,
                    'zones_created': zones_created
                },
                'total_doors': len(all_doors),
                'total_fridges': len(all_fridges),
                'shed_breakdown': shed_details,
                'door_validation': door_validation,
                'fridge_validation': fridge_validation
            }
            
            # Check if we found any doors or fridges
            if len(all_doors) == 0 and len(all_fridges) == 0:
                self.log_test(
                    "Review Request Workflow", 
                    False, 
                    "No doors or fridges found after Excel upload - check if Excel contains blue DOOR cells and yellow FRIDGE cells",
                    json.dumps(results_summary, indent=2)
                )
                return False
            
            self.log_test(
                "Review Request Workflow", 
                True, 
                f"Successfully completed workflow: {len(all_doors)} doors and {len(all_fridges)} fridges detected",
                json.dumps(results_summary, indent=2)
            )
            return True
            
        except Exception as e:
            self.log_test("Review Request Workflow", False, f"Exception: {str(e)}")
            return False

    def test_field_id_matching_issue(self):
        """CRITICAL: Verify Field ID Matching Issue - Execute the specific review request workflow"""
        try:
            print("\n🔍 CRITICAL: VERIFYING FIELD ID MATCHING ISSUE")
            print("=" * 60)
            
            # STEP 1: Get a sample stock intake
            print("STEP 1: Getting sample stock intake...")
            response = self.session.get(f"{self.base_url}/stock-intakes")
            if response.status_code != 200:
                self.log_test("Field ID Matching - Get Stock Intakes", False, f"Failed to get stock intakes, status: {response.status_code}")
                return False
            
            stock_intakes = response.json()
            if not stock_intakes:
                self.log_test("Field ID Matching - Get Stock Intakes", False, "No stock intakes found")
                return False
            
            # Pick the FIRST stock intake
            first_intake = stock_intakes[0]
            intake_field_id = first_intake.get('field_id')
            intake_field_name = first_intake.get('field_name')
            
            print(f"✅ FIRST STOCK INTAKE:")
            print(f"   - field_id: {intake_field_id}")
            print(f"   - field_name: {intake_field_name}")
            
            # STEP 2: Check if that field_id exists in fields
            print(f"\nSTEP 2: Checking if field_id '{intake_field_id}' exists in fields...")
            response = self.session.get(f"{self.base_url}/fields")
            if response.status_code != 200:
                self.log_test("Field ID Matching - Get Fields", False, f"Failed to get fields, status: {response.status_code}")
                return False
            
            fields = response.json()
            if not fields:
                self.log_test("Field ID Matching - Get Fields", False, "No fields found")
                return False
            
            # Search for a field with id matching the stock intake's field_id
            matching_field_by_id = next((f for f in fields if f.get('id') == intake_field_id), None)
            
            if matching_field_by_id:
                print(f"✅ FIELD WITH MATCHING ID FOUND:")
                print(f"   - id: {matching_field_by_id.get('id')}")
                print(f"   - name: {matching_field_by_id.get('name')}")
                actual_field_id = matching_field_by_id.get('id')
            else:
                print(f"❌ NO FIELD WITH ID '{intake_field_id}' FOUND")
                
                # Search for a field with name matching the stock intake's field_name
                print(f"   Searching for field with name '{intake_field_name}'...")
                matching_field_by_name = next((f for f in fields if f.get('name') == intake_field_name), None)
                
                if matching_field_by_name:
                    print(f"✅ FIELD WITH MATCHING NAME FOUND:")
                    print(f"   - id: {matching_field_by_name.get('id')}")
                    print(f"   - name: {matching_field_by_name.get('name')}")
                    actual_field_id = matching_field_by_name.get('id')
                else:
                    print(f"❌ NO FIELD WITH NAME '{intake_field_name}' FOUND EITHER")
                    self.log_test("Field ID Matching", False, f"Neither field_id nor field_name from stock intake found in fields")
                    return False
            
            # STEP 3: Compare IDs
            print(f"\nSTEP 3: Comparing IDs...")
            print(f"   - Stock intake field_id: {intake_field_id}")
            print(f"   - Actual field id with same name: {actual_field_id}")
            
            if intake_field_id == actual_field_id:
                print(f"✅ IDs MATCH - Field ID update logic worked correctly")
                conclusion = "FIELD_IDS_MATCH"
                success = True
                message = "Field IDs match correctly - field_id update logic worked when Excel was re-uploaded"
            else:
                print(f"❌ IDs DO NOT MATCH - Field ID update logic failed")
                conclusion = "FIELD_IDS_MISMATCH"
                success = False
                message = f"Field IDs mismatch - stock intake still references old field ID '{intake_field_id}' but actual field has ID '{actual_field_id}'"
            
            # Additional analysis - check all stock intakes for this pattern
            print(f"\nADDITIONAL ANALYSIS: Checking all stock intakes for field ID mismatches...")
            field_id_map = {f.get('id'): f.get('name') for f in fields}
            field_name_to_id_map = {f.get('name'): f.get('id') for f in fields}
            
            mismatched_intakes = []
            total_intakes = len(stock_intakes)
            
            for intake in stock_intakes:
                intake_fid = intake.get('field_id')
                intake_fname = intake.get('field_name')
                
                # Check if field_id exists in current fields
                if intake_fid not in field_id_map:
                    # Check if field_name exists and has different ID
                    if intake_fname in field_name_to_id_map:
                        correct_id = field_name_to_id_map[intake_fname]
                        mismatched_intakes.append({
                            'intake_id': intake.get('id'),
                            'field_name': intake_fname,
                            'old_field_id': intake_fid,
                            'correct_field_id': correct_id
                        })
            
            print(f"📊 ANALYSIS RESULTS:")
            print(f"   - Total stock intakes: {total_intakes}")
            print(f"   - Intakes with field ID mismatches: {len(mismatched_intakes)}")
            
            if mismatched_intakes:
                print(f"   - Sample mismatched intakes:")
                for i, mismatch in enumerate(mismatched_intakes[:3]):  # Show first 3
                    print(f"     {i+1}. Field '{mismatch['field_name']}': {mismatch['old_field_id']} → {mismatch['correct_field_id']}")
            
            # Final conclusion
            analysis_results = {
                'first_intake_field_id': intake_field_id,
                'first_intake_field_name': intake_field_name,
                'actual_field_id': actual_field_id,
                'ids_match': intake_field_id == actual_field_id,
                'total_stock_intakes': total_intakes,
                'mismatched_intakes_count': len(mismatched_intakes),
                'conclusion': conclusion,
                'sample_mismatches': mismatched_intakes[:5]  # First 5 mismatches
            }
            
            print(f"\n🎯 CRITICAL QUESTION ANSWER:")
            if conclusion == "FIELD_IDS_MATCH":
                print(f"✅ YES - The field_id update logic worked when Excel was re-uploaded")
            else:
                print(f"❌ NO - The field_id update logic did NOT work when Excel was re-uploaded")
                print(f"   Stock intakes are still referencing old deleted field IDs")
            
            self.log_test(
                "Field ID Matching Issue", 
                success, 
                message,
                json.dumps(analysis_results, indent=2)
            )
            return success
            
        except Exception as e:
            self.log_test("Field ID Matching Issue", False, f"Exception: {str(e)}")
            return False

    def investigate_lost_stock_data(self):
        """URGENT: Investigate Lost Stock Data - Execute the specific review request workflow"""
        try:
            print("\n🚨 URGENT: INVESTIGATING LOST STOCK DATA")
            print("=" * 60)
            
            investigation_results = {}
            
            # STEP 1: Check if stock intakes still exist
            print("STEP 1: Checking stock intakes...")
            response = self.session.get(f"{self.base_url}/stock-intakes")
            if response.status_code != 200:
                self.log_test("Stock Intakes Check", False, f"Failed to get stock intakes, status: {response.status_code}")
                return False
            
            stock_intakes = response.json()
            investigation_results['stock_intakes_count'] = len(stock_intakes)
            
            print(f"📊 STOCK INTAKES: {len(stock_intakes)} found")
            
            if stock_intakes:
                # Show sample intake with field_id, zone_id, shed_id
                sample_intake = stock_intakes[0]
                investigation_results['sample_stock_intake'] = {
                    'id': sample_intake.get('id'),
                    'field_id': sample_intake.get('field_id'),
                    'zone_id': sample_intake.get('zone_id'),
                    'shed_id': sample_intake.get('shed_id'),
                    'quantity': sample_intake.get('quantity'),
                    'field_name': sample_intake.get('field_name')
                }
                print(f"📋 SAMPLE INTAKE:")
                print(f"   - ID: {sample_intake.get('id')}")
                print(f"   - Field ID: {sample_intake.get('field_id')}")
                print(f"   - Zone ID: {sample_intake.get('zone_id')}")
                print(f"   - Shed ID: {sample_intake.get('shed_id')}")
                print(f"   - Quantity: {sample_intake.get('quantity')}")
                print(f"   - Field Name: {sample_intake.get('field_name')}")
            else:
                print("⚠️  NO STOCK INTAKES FOUND!")
                investigation_results['sample_stock_intake'] = None
            
            # STEP 2: Check fields
            print("\nSTEP 2: Checking fields...")
            response = self.session.get(f"{self.base_url}/fields")
            if response.status_code != 200:
                self.log_test("Fields Check", False, f"Failed to get fields, status: {response.status_code}")
                return False
            
            fields = response.json()
            investigation_results['fields_count'] = len(fields)
            
            print(f"📊 FIELDS: {len(fields)} found")
            
            if fields:
                sample_field = fields[0]
                investigation_results['sample_field'] = {
                    'id': sample_field.get('id'),
                    'name': sample_field.get('name'),
                    'crop_type': sample_field.get('crop_type'),
                    'variety': sample_field.get('variety')
                }
                print(f"📋 SAMPLE FIELD:")
                print(f"   - ID: {sample_field.get('id')}")
                print(f"   - Name: {sample_field.get('name')}")
                print(f"   - Crop Type: {sample_field.get('crop_type')}")
                print(f"   - Variety: {sample_field.get('variety')}")
            else:
                print("⚠️  NO FIELDS FOUND!")
                investigation_results['sample_field'] = None
            
            # STEP 3: Check zones
            print("\nSTEP 3: Checking zones...")
            response = self.session.get(f"{self.base_url}/zones")
            if response.status_code != 200:
                self.log_test("Zones Check", False, f"Failed to get zones, status: {response.status_code}")
                return False
            
            zones = response.json()
            investigation_results['zones_count'] = len(zones)
            
            print(f"📊 ZONES: {len(zones)} found")
            
            if zones:
                sample_zone = zones[0]
                investigation_results['sample_zone'] = {
                    'id': sample_zone.get('id'),
                    'name': sample_zone.get('name'),
                    'shed_id': sample_zone.get('shed_id'),
                    'total_quantity': sample_zone.get('total_quantity')
                }
                print(f"📋 SAMPLE ZONE:")
                print(f"   - ID: {sample_zone.get('id')}")
                print(f"   - Name: {sample_zone.get('name')}")
                print(f"   - Shed ID: {sample_zone.get('shed_id')}")
                print(f"   - Total Quantity: {sample_zone.get('total_quantity')}")
            else:
                print("⚠️  NO ZONES FOUND!")
                investigation_results['sample_zone'] = None
            
            # STEP 4: Check sheds
            print("\nSTEP 4: Checking sheds...")
            response = self.session.get(f"{self.base_url}/sheds")
            if response.status_code != 200:
                self.log_test("Sheds Check", False, f"Failed to get sheds, status: {response.status_code}")
                return False
            
            sheds = response.json()
            investigation_results['sheds_count'] = len(sheds)
            
            print(f"📊 SHEDS: {len(sheds)} found")
            
            if sheds:
                sample_shed = sheds[0]
                investigation_results['sample_shed'] = {
                    'id': sample_shed.get('id'),
                    'name': sample_shed.get('name'),
                    'width': sample_shed.get('width'),
                    'height': sample_shed.get('height')
                }
                print(f"📋 SAMPLE SHED:")
                print(f"   - ID: {sample_shed.get('id')}")
                print(f"   - Name: {sample_shed.get('name')}")
                print(f"   - Width: {sample_shed.get('width')}")
                print(f"   - Height: {sample_shed.get('height')}")
            else:
                print("⚠️  NO SHEDS FOUND!")
                investigation_results['sample_shed'] = None
            
            # STEP 5: Cross-reference IDs
            print("\nSTEP 5: Cross-referencing IDs...")
            
            if stock_intakes:
                # Create ID sets for comparison
                field_ids = set([f.get('id') for f in fields])
                zone_ids = set([z.get('id') for z in zones])
                shed_ids = set([s.get('id') for s in sheds])
                
                # Check for orphaned stock intakes
                orphaned_field_ids = []
                orphaned_zone_ids = []
                orphaned_shed_ids = []
                
                for intake in stock_intakes:
                    intake_field_id = intake.get('field_id')
                    intake_zone_id = intake.get('zone_id')
                    intake_shed_id = intake.get('shed_id')
                    
                    if intake_field_id not in field_ids:
                        orphaned_field_ids.append(intake_field_id)
                    
                    if intake_zone_id not in zone_ids:
                        orphaned_zone_ids.append(intake_zone_id)
                    
                    if intake_shed_id not in shed_ids:
                        orphaned_shed_ids.append(intake_shed_id)
                
                # Remove duplicates
                orphaned_field_ids = list(set(orphaned_field_ids))
                orphaned_zone_ids = list(set(orphaned_zone_ids))
                orphaned_shed_ids = list(set(orphaned_shed_ids))
                
                investigation_results['orphaned_analysis'] = {
                    'orphaned_field_ids': orphaned_field_ids,
                    'orphaned_zone_ids': orphaned_zone_ids,
                    'orphaned_shed_ids': orphaned_shed_ids
                }
                
                print(f"🔍 ID CROSS-REFERENCE ANALYSIS:")
                print(f"   - Stock intakes with orphaned field_ids: {len(orphaned_field_ids)}")
                print(f"   - Stock intakes with orphaned zone_ids: {len(orphaned_zone_ids)}")
                print(f"   - Stock intakes with orphaned shed_ids: {len(orphaned_shed_ids)}")
                
                if orphaned_field_ids:
                    print(f"   - Orphaned field IDs: {orphaned_field_ids[:5]}{'...' if len(orphaned_field_ids) > 5 else ''}")
                
                if orphaned_zone_ids:
                    print(f"   - Orphaned zone IDs: {orphaned_zone_ids[:5]}{'...' if len(orphaned_zone_ids) > 5 else ''}")
                
                if orphaned_shed_ids:
                    print(f"   - Orphaned shed IDs: {orphaned_shed_ids[:5]}{'...' if len(orphaned_shed_ids) > 5 else ''}")
                
                # Determine the issue
                if orphaned_field_ids or orphaned_zone_ids or orphaned_shed_ids:
                    investigation_results['conclusion'] = "ORPHANED_DATA"
                    print(f"\n🚨 CONCLUSION: Stock intakes are ORPHANED (referencing old IDs that no longer exist)")
                else:
                    investigation_results['conclusion'] = "DATA_INTACT"
                    print(f"\n✅ CONCLUSION: Stock intakes are intact and properly referenced")
            else:
                investigation_results['orphaned_analysis'] = None
                investigation_results['conclusion'] = "NO_STOCK_INTAKES"
                print(f"\n🚨 CONCLUSION: Stock intakes were DELETED (no stock intakes found)")
            
            # Final summary
            print(f"\n📋 INVESTIGATION SUMMARY:")
            print(f"   - Stock Intakes: {investigation_results['stock_intakes_count']}")
            print(f"   - Fields: {investigation_results['fields_count']}")
            print(f"   - Zones: {investigation_results['zones_count']}")
            print(f"   - Sheds: {investigation_results['sheds_count']}")
            print(f"   - Conclusion: {investigation_results['conclusion']}")
            
            self.log_test(
                "Lost Stock Data Investigation", 
                True, 
                f"Investigation completed: {investigation_results['conclusion']}",
                json.dumps(investigation_results, indent=2)
            )
            return True
            
        except Exception as e:
            self.log_test("Lost Stock Data Investigation", False, f"Exception: {str(e)}")
            return False

    def run_all_tests(self):
        """Run all tests in sequence"""
        print(f"🧪 Starting Stock Control Backend API Tests")
        print(f"🔗 Backend URL: {self.base_url}")
        print("=" * 60)
        
        # Test sequence - prioritizing field ID matching issue
        tests = [
            ("API Health Check", self.test_api_health),
            ("CRITICAL: Field ID Matching Issue", self.test_field_id_matching_issue),
            ("URGENT: Lost Stock Data Investigation", self.investigate_lost_stock_data),
            ("Review Request Workflow", self.test_review_request_workflow),
            ("Clear All Data with Fridges", self.test_clear_data_with_fridges),
            ("Fridge API Endpoints", self.test_fridge_api_endpoints),
            ("Excel Fridge Parsing", self.test_excel_fridge_parsing),
            ("Hardcoded Admin Login", self.test_hardcoded_admin_login),
            ("Type Column Integration", self.test_type_column_integration),
            ("Fields API and Grade Assignment", self.test_fields_api_grades),
            ("Shed CRUD Operations", self.test_shed_crud),
            ("Zone CRUD Operations", self.test_zone_crud),
            ("Stock Intake with Grade", self.test_stock_intake_with_grade)
        ]
        
        passed = 0
        total = len(tests)
        
        for test_name, test_func in tests:
            print(f"\n🔍 Running: {test_name}")
            if test_func():
                passed += 1
        
        print("\n" + "=" * 60)
        print(f"📊 Test Results: {passed}/{total} tests passed")
        
        # Show failed tests
        failed_tests = [result for result in self.test_results if not result['success']]
        if failed_tests:
            print("\n❌ Failed Tests:")
            for test in failed_tests:
                print(f"   • {test['test']}: {test['message']}")
                if test['details']:
                    print(f"     Details: {test['details']}")
        
        return passed == total

if __name__ == "__main__":
    tester = StockControlTester()
    success = tester.run_all_tests()
    exit(0 if success else 1)